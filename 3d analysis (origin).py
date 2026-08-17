#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import glob
import numpy as np
from tkinter import Tk, filedialog, simpledialog, messagebox
from skimage import io
from skimage.measure import label, regionprops
from scipy.ndimage import distance_transform_edt
from numpy.polynomial import polynomial as poly
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


NI_LO = 2.99
NI_HI = 3.84


def extract_registered_number_from_star_folder(name: str):
    if "★" not in name:
        return None
    m = re.search(r"3D_(\d+)", name)
    return int(m.group(1)) if m else None


def normalize_mask_to_shape(mask_2d, target_shape):
    out = np.zeros(target_shape, dtype=bool)
    h = min(mask_2d.shape[0], target_shape[0])
    w = min(mask_2d.shape[1], target_shape[1])
    out[:h, :w] = mask_2d[:h, :w]
    return out


def normalize_array_to_shape(arr_2d, target_shape, fill_value=np.nan):
    out = np.full(target_shape, fill_value, dtype=arr_2d.dtype)
    h = min(arr_2d.shape[0], target_shape[0])
    w = min(arr_2d.shape[1], target_shape[1])
    out[:h, :w] = arr_2d[:h, :w]
    return out


def ask_analysis_mode():
    root = Tk()
    root.withdraw()
    yes_phase_sep = messagebox.askyesno(
        "분석 모드 선택",
        "상분리 분석 모드로 진행할까요?\n\n"
        "YES = 상분리 분석 모드\n"
        "NO = 완충/프리스틴 분석 모드"
    )
    root.destroy()
    return "phase_separation" if yes_phase_sep else "pristine_fullcharge"


def ask_voxel_size_um():
    s = simpledialog.askstring(
        "voxel size (µm)",
        "voxel edge length를 µm 단위로 입력하세요.\n예: 60 nm -> 0.06",
        initialvalue="0.06"
    )
    if not s:
        return 0.06
    v = float(s)
    if v <= 0:
        raise ValueError("❌ voxel size 입력이 올바르지 않습니다.")
    return v


def ask_min_domain_voxels_after_intersection():
    s = simpledialog.askstring(
        "도메인 최소 복셀 수",
        "H2/H3 phase domain을 shell/bulk 중 더 많이 겹치는 쪽으로 통째로 배정하기 전에,\n"
        "최종 domain 통계에서 유지할 최소 connected voxel 수를 입력하세요.\n\n"
        "예: 3 입력 -> H2/H3 phase domain 자체가 1~2 voxel이면 제거",
        initialvalue="3"
    )
    if not s:
        return 3
    v = int(float(s))
    if v < 1:
        raise ValueError("❌ 도메인 최소 복셀 수 입력이 올바르지 않습니다.")
    return v


def ask_peak_base_and_margin():
    peakref_str = simpledialog.askstring(
        "기준 peak_site 범위",
        "평균 계산용 기준 범위 [low,high]를 입력하세요.\n예: 8353.6,8356.2",
        initialvalue="8353.6,8356.2"
    )
    try:
        a0, b0 = peakref_str.replace(" ", "").split(",")
        peakref = [float(a0), float(b0)]
        if peakref[0] > peakref[1]:
            peakref.sort()
    except Exception:
        peakref = [8353.6, 8356.2]

    margin_str = simpledialog.askstring(
        "상/하단 마진 (eV)",
        "유효 peak_site 범위 = [low-low_margin, high+high_margin]\n"
        "형식: low_margin,high_margin\n예: 0,0  또는  0.5,0.5",
        initialvalue="0,0"
    )
    try:
        lm_s, hm_s = margin_str.replace(" ", "").split(",")
        low_margin = max(float(lm_s), 0.0)
        high_margin = max(float(hm_s), 0.0)
    except Exception:
        low_margin, high_margin = 0.0, 0.0

    return peakref, low_margin, high_margin


def ask_single_bnl_profile(profile_name):
    messagebox.showinfo('BNL 설정', f'{profile_name} 설정을 입력하세요.')

    energy_file = filedialog.askopenfilename(title=f'{profile_name} energy.txt 선택')
    if not energy_file:
        raise ValueError(f'❌ {profile_name} energy.txt가 선택되지 않았습니다.')
    eng = np.loadtxt(energy_file)

    ev_step = float(simpledialog.askstring(
        f'{profile_name} ev_step',
        f'{profile_name} ev_step(eV) 입력',
        initialvalue='1'
    ) or '1')

    energy_axis_reverse = messagebox.askyesno(
        f'{profile_name} Energy axis reverse?',
        f'{profile_name}: img_stack을 [::-1,:,:]로 뒤집을까요?'
    )

    soc_str = simpledialog.askstring(
        f"{profile_name} SOC 기준 에너지",
        f"{profile_name} soc0_energy,soc100_energy\n예: 8353.6,8356.2",
        initialvalue="8353.6,8356.2"
    )
    try:
        a, b = soc_str.replace(" ", "").split(",")
        soc0_energy = float(a)
        soc100_energy = float(b)
    except Exception:
        raise ValueError(f"❌ {profile_name} SOC 기준 에너지 입력 오류")

    return {
        'eng': np.asarray(eng, dtype=float).ravel(),
        'ev_step': float(ev_step),
        'energy_axis_reverse': bool(energy_axis_reverse),
        'soc0_energy': float(soc0_energy),
        'soc100_energy': float(soc100_energy),
        'ni_peak_lo': float(soc0_energy),
        'ni_peak_hi': float(soc100_energy),
    }


def ask_bnl_profiles():
    return {
        'BNL1': ask_single_bnl_profile('BNL1'),
        'BNL2': ask_single_bnl_profile('BNL2')
    }


def ask_phase_fraction_excel():
    use_excel = messagebox.askyesno(
        "in situ XRD H2/H3 상분율 엑셀",
        "in situ XRD 기반 SOC-H2/H3 fraction 엑셀을 적용할까요?\n\n"
        "YES = 엑셀 선택\n"
        "NO = h2_h3 split energy는 NaN 저장"
    )

    if not use_excel:
        return None

    path = filedialog.askopenfilename(
        title="in situ XRD H2/H3 fraction Excel 선택",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )

    if not path:
        return None

    return load_phase_fraction_excel(path)


def _find_column(headers, candidates):
    headers_lower = [str(h).strip().lower() if h is not None else "" for h in headers]
    for cand in candidates:
        cand = cand.lower()
        for i, h in enumerate(headers_lower):
            if cand in h:
                return i
    return None


def load_phase_fraction_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("❌ in situ XRD 상분율 엑셀에 데이터가 부족합니다.")

    header_row_idx = None
    soc_col = None
    h3_col = None
    h2_col = None

    for ridx, row in enumerate(rows[:20]):
        headers = list(row)
        soc_col = _find_column(headers, ["soc", "state of charge", "charge"])
        h3_col = _find_column(headers, ["h3"])
        h2_col = _find_column(headers, ["h2"])
        if soc_col is not None and (h3_col is not None or h2_col is not None):
            header_row_idx = ridx
            break

    if header_row_idx is None:
        raise ValueError(
            "❌ 엑셀에서 SOC 컬럼과 H3 또는 H2 fraction 컬럼을 찾지 못했습니다.\n"
            "컬럼명에 SOC, H3 또는 H2가 포함되게 해주세요."
        )

    soc_list = []
    h3_frac_list = []

    for row in rows[header_row_idx + 1:]:
        try:
            soc = row[soc_col]
            if soc is None:
                continue
            soc = float(soc)

            if h3_col is not None:
                h3f = row[h3_col]
                if h3f is None:
                    continue
                h3f = float(h3f)
            else:
                h2f = row[h2_col]
                if h2f is None:
                    continue
                h3f = 1.0 - float(h2f)

            if soc > 1.5:
                soc = soc / 100.0

            if h3f > 1.5:
                h3f = h3f / 100.0

            if np.isfinite(soc) and np.isfinite(h3f):
                soc_list.append(soc)
                h3_frac_list.append(np.clip(h3f, 0.0, 1.0))
        except Exception:
            continue

    if len(soc_list) < 2:
        raise ValueError("❌ SOC-H3 fraction 유효 데이터가 2개 미만입니다.")

    soc_arr = np.asarray(soc_list, dtype=float)
    h3_arr = np.asarray(h3_frac_list, dtype=float)

    order = np.argsort(soc_arr)
    soc_arr = soc_arr[order]
    h3_arr = h3_arr[order]

    return {
        "path": path,
        "soc": soc_arr,
        "h3_fraction": h3_arr,
    }


def interpolate_h3_fraction_from_soc(soc, phase_fraction_cfg):
    if phase_fraction_cfg is None or not np.isfinite(soc):
        return np.nan

    soc_axis = phase_fraction_cfg["soc"]
    h3_axis = phase_fraction_cfg["h3_fraction"]

    soc_c = float(np.clip(soc, np.nanmin(soc_axis), np.nanmax(soc_axis)))
    return float(np.interp(soc_c, soc_axis, h3_axis))


def compute_h2_h3_split_energy_from_fraction(peak_values, target_h3_fraction):
    vals = np.asarray(peak_values, dtype=float)
    vals = vals[np.isfinite(vals)]

    if vals.size == 0:
        return np.nan

    if not np.isfinite(target_h3_fraction):
        return np.nan

    f_h3 = float(np.clip(target_h3_fraction, 0.0, 1.0))

    if f_h3 <= 0:
        return float(np.nanmax(vals))
    if f_h3 >= 1:
        return float(np.nanmin(vals))

    percentile = 100.0 * (1.0 - f_h3)
    return float(np.nanpercentile(vals, percentile))


def peak_to_soc(peak, soc0_energy, soc100_energy):
    if soc100_energy == soc0_energy:
        return np.nan
    return float((peak - soc0_energy) / (soc100_energy - soc0_energy))


def ask_parent_directories_with_profiles(profiles):
    parents = []
    while True:
        folder = filedialog.askdirectory(title='상위 폴더 선택 (취소 누르면 선택 종료)')
        if not folder:
            break

        yes_bnl1 = messagebox.askyesno(
            'BNL 프로필 선택',
            '이 상위 폴더를 BNL1로 처리할까요?\n\nYES = BNL1\nNO = BNL2'
        )
        profile_name = 'BNL1' if yes_bnl1 else 'BNL2'

        parents.append({'folder': folder, 'profile_name': profile_name})

        more = messagebox.askyesno('폴더 추가 선택', '다른 상위 폴더를 추가로 선택하시겠습니까?')
        if not more:
            break

    return parents


def ask_output_excel_path():
    path = filedialog.asksaveasfilename(
        title="최종 엑셀 저장 경로 선택",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not path:
        raise ValueError("❌ 저장할 엑셀 경로를 선택하지 않았습니다.")
    return path


def find_star_folders_recursively(parent_dir):
    out = []
    pat = re.compile(r"^3D_(\d+)_fit_3★$")
    for cur_root, dirnames, _ in os.walk(parent_dir):
        for d in dirnames:
            if pat.match(d):
                out.append((cur_root, d))
    return sorted(out, key=lambda x: (x[0], x[1]))


def parse_slice_number(path_or_name):
    name = os.path.basename(path_or_name)
    m = re.search(r"_(\d+)\.npy$", name)
    return int(m.group(1)) if m else None


def peak_to_ni(peak, peak_lo, peak_hi):
    peak = np.asarray(peak, dtype=float)
    peak_lo = float(peak_lo)
    peak_hi = float(peak_hi)
    if peak_hi == peak_lo:
        raise ValueError("peak_to_ni mapping range is invalid")
    peak_c = np.clip(peak, peak_lo, peak_hi)
    ni = NI_LO + (peak_c - peak_lo) / (peak_hi - peak_lo) * (NI_HI - NI_LO)
    return np.clip(ni, NI_LO, NI_HI)


def filter_and_clip_peak_values(values, low, high, low_margin, high_margin):
    values = np.asarray(values, dtype=float)
    values = values[np.isfinite(values)]

    eff_low = float(low) - float(low_margin)
    eff_high = float(high) + float(high_margin)

    valid_values = values[(values >= eff_low) & (values <= eff_high)]
    if valid_values.size == 0:
        return np.array([], dtype=float)

    return np.clip(valid_values, float(low), float(high))


def align_energy_axis_for_legacy_fit(img_stack, eng, energy_axis_reverse=True):
    eng = np.array(eng, dtype=float).ravel()

    if img_stack.ndim != 3:
        raise ValueError(f"img_stack must be 3D (E,H,W). Got {img_stack.shape}")
    if img_stack.shape[0] != len(eng):
        raise ValueError(f"energy length mismatch: E={img_stack.shape[0]} vs len(eng)={len(eng)}")

    if energy_axis_reverse:
        img_stack = img_stack[::-1, :, :]

    if eng[0] > eng[-1]:
        eng = eng[::-1]

    diffs = np.diff(eng)
    if np.any(diffs <= 0):
        raise ValueError("eng is not strictly increasing. energy.txt order may be wrong.")

    return img_stack, eng


def polynomial_second_fit_separate(img_stack, eng, fit_num, ev_step):
    y = img_stack.reshape(len(img_stack), -1)
    y_new = np.zeros((fit_num * 2 + 1, y.shape[1]), dtype=np.float32)
    mp_list = []

    for i in range(y.shape[1]):
        a = int(np.argmax(y[:, i]))
        if a + fit_num + 1 > len(y):
            y_new[:, i] = y[len(y) - 2 * fit_num - 1:len(y) + 1, i]
        elif a - fit_num < 0:
            y_new[:, i] = y[0: 2 * fit_num + 1, i]
        else:
            y_new[:, i] = y[a - fit_num:a + fit_num + 1, i]
        mp_list.append(a)

    x = np.linspace(0, ev_step * (len(y_new) - 1), len(y_new))
    coefs, _ = poly.polyfit(x, y_new, 2, full=True)

    ind = [a - fit_num for a in mp_list]
    denom = 2.0 * coefs[2]
    with np.errstate(divide="ignore", invalid="ignore"):
        x0 = -(coefs[1] / denom) + eng[ind]
    x0[~np.isfinite(x0)] = np.nan

    return x0.reshape(img_stack.shape[1], img_stack.shape[2])


def largest_domain_width_vox(domain_mask_3d):
    if not np.any(domain_mask_3d):
        return np.nan
    edt = distance_transform_edt(domain_mask_3d)
    vals = edt[domain_mask_3d]
    vals = vals[np.isfinite(vals)]
    if vals.size == 0:
        return np.nan
    return float(2.0 * np.max(vals))


def find_mask_subdir_by_tag(mask_root, tag):
    cands = []
    for name in os.listdir(mask_root):
        full = os.path.join(mask_root, name)
        if not os.path.isdir(full):
            continue
        if tag == "h3" and "-1(" in name:
            cands.append(full)
        elif tag == "h2" and "-2(" in name:
            cands.append(full)
        elif tag == "shell" and "-3(" in name:
            cands.append(full)
        elif tag == "bulk" and "-4(" in name:
            cands.append(full)
    cands.sort()
    return cands[0] if cands else None


def load_mask_stack(mask_dir, slice_numbers, target_shape):
    masks = []
    for s in slice_numbers:
        if mask_dir is None:
            masks.append(np.zeros(target_shape, dtype=bool))
            continue
        path = os.path.join(mask_dir, f"filtered_mask_slice_{s:04d}.npy")
        if not os.path.exists(path):
            masks.append(np.zeros(target_shape, dtype=bool))
            continue
        m = np.load(path).astype(bool)
        if m.shape != target_shape:
            m = normalize_mask_to_shape(m, target_shape)
        masks.append(m)
    return np.stack(masks, axis=0).astype(bool)


def assign_phase_domains_to_shell_or_bulk(phase_mask_3d, shell_mask_3d, bulk_mask_3d,
                                          min_voxels=3, connectivity=1):
    phase_mask_3d = np.asarray(phase_mask_3d, dtype=bool)
    shell_mask_3d = np.asarray(shell_mask_3d, dtype=bool)
    bulk_mask_3d = np.asarray(bulk_mask_3d, dtype=bool)

    shell_assigned = np.zeros_like(phase_mask_3d, dtype=bool)
    bulk_assigned = np.zeros_like(phase_mask_3d, dtype=bool)

    if phase_mask_3d.ndim != 3 or not np.any(phase_mask_3d):
        return shell_assigned, bulk_assigned

    min_voxels = max(1, int(min_voxels))

    lbl = label(phase_mask_3d.astype(np.uint8), connectivity=connectivity)
    if lbl.max() == 0:
        return shell_assigned, bulk_assigned

    for lab in range(1, int(lbl.max()) + 1):
        dom = (lbl == lab)
        dom_voxels = int(np.count_nonzero(dom))

        if dom_voxels < min_voxels:
            continue

        shell_count = int(np.count_nonzero(dom & shell_mask_3d))
        bulk_count = int(np.count_nonzero(dom & bulk_mask_3d))

        if shell_count == 0 and bulk_count == 0:
            continue

        if bulk_count > shell_count:
            bulk_assigned[dom] = True
        else:
            shell_assigned[dom] = True

    return shell_assigned, bulk_assigned


def analyze_domains_metric(domain_mask_3d, peak_3d, voxel_um, low, high, low_margin, high_margin, ni_peak_lo, ni_peak_hi):
    labeled = label(domain_mask_3d.astype(np.uint8), connectivity=1)
    props = regionprops(labeled)

    mean_ni_list = []
    width_um_list = []
    volume_um3_list = []

    for p in props:
        dom_mask = (labeled == p.label)
        vals = filter_and_clip_peak_values(
            peak_3d[dom_mask], low, high, low_margin, high_margin
        )

        if vals.size == 0:
            dom_mean_ni = np.nan
        else:
            dom_ni = peak_to_ni(vals, ni_peak_lo, ni_peak_hi)
            dom_mean_ni = float(np.mean(dom_ni))

        width_vox = largest_domain_width_vox(dom_mask)
        width_um = float(width_vox * voxel_um) if np.isfinite(width_vox) else np.nan
        volume_um3 = float(int(np.count_nonzero(dom_mask)) * (voxel_um ** 3))

        mean_ni_list.append(dom_mean_ni)
        width_um_list.append(width_um)
        volume_um3_list.append(volume_um3)

    return {
        "count": int(len(props)),
        "mean_ni_list": mean_ni_list,
        "width_um_list": width_um_list,
        "volume_um3_list": volume_um3_list,
    }


def analyze_voxel_peak_energy(mask_3d, peak_3d, low, high, low_margin, high_margin):
    vals = filter_and_clip_peak_values(
        peak_3d[mask_3d], low, high, low_margin, high_margin
    )
    if vals.size == 0:
        return []
    return vals.astype(float).tolist()


def build_summary_only_result(reg_dir, effective_diameter_um, particle_mean_ni, particle_std_ni,
                              h2_h3_split_energy=np.nan, particle_soc=np.nan, xrd_h3_fraction=np.nan):
    return {
        "particle_folder_path": reg_dir,
        "particle_name": os.path.basename(os.path.dirname(reg_dir)) + f"_registered{os.path.basename(reg_dir).replace('registered', '')}",
        "effective_diameter_um": effective_diameter_um,
        "particle_mean_ni": particle_mean_ni,
        "particle_std_ni": particle_std_ni,
        "hist_bulk_voxel_energy": [],
        "hist_shell_voxel_energy": [],
        "bulk_h3_width": [],
        "shell_h3_width": [],
        "bulk_h2_width": [],
        "shell_h2_width": [],
        "bulk_h3_volume": [],
        "shell_h3_volume": [],
        "bulk_h2_volume": [],
        "shell_h2_volume": [],
        "summary_shell_h3_count": np.nan,
        "summary_bulk_h3_count": np.nan,
        "summary_shell_h2_count": np.nan,
        "summary_bulk_h2_count": np.nan,
        "summary_total_h3_count": np.nan,
        "summary_total_h2_count": np.nan,
        "summary_shell_h3_fraction": np.nan,
        "summary_bulk_h3_fraction": np.nan,
        "summary_shell_avg_ni": np.nan,
        "summary_bulk_avg_ni": np.nan,
        "min_domain_voxels_after_intersection": np.nan,
        "h2_peaksite_values": [],
        "h3_peaksite_values": [],
        "h2_peaksite_mean": np.nan,
        "h3_peaksite_mean": np.nan,
        "h2_h3_split_energy": h2_h3_split_energy,
        "particle_soc": particle_soc,
        "xrd_h3_fraction": xrd_h3_fraction,
        "reg_dir": reg_dir,
        "roi_mask_dir": None,
        "h3_dir": None,
        "h2_dir": None,
        "shell_dir": None,
        "bulk_dir": None,
    }


def process_registered_particle(reg_dir, roi_mask_dir, profiles_cfg, phase_fraction_cfg,
                                peakref, low_margin, high_margin, voxel_um,
                                analysis_mode, min_domain_voxels_after_intersection):
    tiff_files = sorted(glob.glob(os.path.join(reg_dir, "*.tif")) + glob.glob(os.path.join(reg_dir, "*.tiff")))
    if not tiff_files:
        raise ValueError(f"No tif/tiff in folder: {reg_dir}")

    roi_files = sorted(glob.glob(os.path.join(roi_mask_dir, "roi_mask_slice_*.npy")))
    if len(roi_files) == 0:
        raise ValueError(f"ROI mask 없음: {roi_mask_dir}")

    eng = profiles_cfg["eng"]
    ev_step = profiles_cfg["ev_step"]
    energy_axis_reverse = profiles_cfg["energy_axis_reverse"]
    ni_peak_lo = profiles_cfg["ni_peak_lo"]
    ni_peak_hi = profiles_cfg["ni_peak_hi"]
    soc0_energy = profiles_cfg["soc0_energy"]
    soc100_energy = profiles_cfg["soc100_energy"]

    raw_particle_masks = []
    raw_peak_maps = []
    slice_shapes = []
    slice_numbers = []

    fit_num = 3

    for roi_path in roi_files:
        s = parse_slice_number(roi_path)
        if s is None:
            continue
        tif_idx = s - 1
        if tif_idx < 0 or tif_idx >= len(tiff_files):
            continue

        roi_mask = np.load(roi_path).astype(bool)
        img_stack_raw = io.imread(tiff_files[tif_idx])
        img_stack, eng_aligned = align_energy_axis_for_legacy_fit(
            img_stack_raw, eng, energy_axis_reverse=energy_axis_reverse
        )
        peak_map = polynomial_second_fit_separate(
            img_stack, eng_aligned, fit_num=fit_num, ev_step=ev_step
        )

        target_shape = peak_map.shape
        if roi_mask.shape != target_shape:
            roi_mask = normalize_mask_to_shape(roi_mask, target_shape)

        raw_particle_masks.append(roi_mask)
        raw_peak_maps.append(peak_map)
        slice_shapes.append(peak_map.shape)
        slice_numbers.append(s)

    if len(raw_peak_maps) == 0:
        raise ValueError("분석 가능한 slice가 없습니다.")

    min_h = min(s[0] for s in slice_shapes)
    min_w = min(s[1] for s in slice_shapes)
    common_shape = (min_h, min_w)

    particle_mask_slices = []
    peak_slices = []

    for roi_mask, peak_map in zip(raw_particle_masks, raw_peak_maps):
        if roi_mask.shape != common_shape:
            roi_mask = normalize_mask_to_shape(roi_mask, common_shape)
        if peak_map.shape != common_shape:
            peak_map = normalize_array_to_shape(peak_map, common_shape, fill_value=np.nan)

        particle_mask_slices.append(roi_mask)
        peak_slices.append(peak_map)

    particle_mask_3d = np.stack(particle_mask_slices, axis=0)
    peak_3d = np.stack(peak_slices, axis=0)

    particle_vals = filter_and_clip_peak_values(
        peak_3d[particle_mask_3d], peakref[0], peakref[1], low_margin, high_margin
    )
    if particle_vals.size == 0:
        raise ValueError("ROI 안 유효 peak 값이 없습니다.")

    particle_ni = peak_to_ni(particle_vals, ni_peak_lo, ni_peak_hi)
    particle_mean_ni = float(np.mean(particle_ni))
    particle_std_ni = float(np.std(particle_ni))

    particle_mean_peak = float(np.mean(particle_vals))
    particle_soc = peak_to_soc(particle_mean_peak, soc0_energy, soc100_energy)
    xrd_h3_fraction = interpolate_h3_fraction_from_soc(particle_soc, phase_fraction_cfg)
    h2_h3_split_energy = compute_h2_h3_split_energy_from_fraction(particle_vals, xrd_h3_fraction)

    voxel_count_particle = int(np.count_nonzero(particle_mask_3d))
    total_vol_um3 = voxel_count_particle * (voxel_um ** 3)
    effective_diameter_um = float(2.0 * (((3.0 * total_vol_um3) / (4.0 * np.pi)) ** (1.0 / 3.0)))

    if analysis_mode == "pristine_fullcharge":
        return build_summary_only_result(
            reg_dir=reg_dir,
            effective_diameter_um=effective_diameter_um,
            particle_mean_ni=particle_mean_ni,
            particle_std_ni=particle_std_ni,
            h2_h3_split_energy=h2_h3_split_energy,
            particle_soc=particle_soc,
            xrd_h3_fraction=xrd_h3_fraction
        )

    h3_dir = find_mask_subdir_by_tag(roi_mask_dir, "h3")
    h2_dir = find_mask_subdir_by_tag(roi_mask_dir, "h2")
    shell_dir = find_mask_subdir_by_tag(roi_mask_dir, "shell")
    bulk_dir = find_mask_subdir_by_tag(roi_mask_dir, "bulk")

    missing_mask = any(x is None for x in [h3_dir, h2_dir, shell_dir, bulk_dir])
    if missing_mask:
        return build_summary_only_result(
            reg_dir=reg_dir,
            effective_diameter_um=effective_diameter_um,
            particle_mean_ni=particle_mean_ni,
            particle_std_ni=particle_std_ni,
            h2_h3_split_energy=h2_h3_split_energy,
            particle_soc=particle_soc,
            xrd_h3_fraction=xrd_h3_fraction
        )

    h3_mask_3d = load_mask_stack(h3_dir, slice_numbers, common_shape) & particle_mask_3d
    h2_mask_3d = load_mask_stack(h2_dir, slice_numbers, common_shape) & particle_mask_3d
    shell_mask_3d = load_mask_stack(shell_dir, slice_numbers, common_shape) & particle_mask_3d
    bulk_mask_3d = load_mask_stack(bulk_dir, slice_numbers, common_shape) & particle_mask_3d

    h2_peaksite_values = analyze_voxel_peak_energy(
        h2_mask_3d, peak_3d, peakref[0], peakref[1], low_margin, high_margin
    )
    h3_peaksite_values = analyze_voxel_peak_energy(
        h3_mask_3d, peak_3d, peakref[0], peakref[1], low_margin, high_margin
    )

    h2_peaksite_mean = float(np.mean(h2_peaksite_values)) if len(h2_peaksite_values) > 0 else np.nan
    h3_peaksite_mean = float(np.mean(h3_peaksite_values)) if len(h3_peaksite_values) > 0 else np.nan

    shell_only = shell_mask_3d
    bulk_only = bulk_mask_3d

    shell_h3, bulk_h3 = assign_phase_domains_to_shell_or_bulk(
        h3_mask_3d,
        shell_mask_3d,
        bulk_mask_3d,
        min_voxels=min_domain_voxels_after_intersection,
        connectivity=1
    )
    shell_h2, bulk_h2 = assign_phase_domains_to_shell_or_bulk(
        h2_mask_3d,
        shell_mask_3d,
        bulk_mask_3d,
        min_voxels=min_domain_voxels_after_intersection,
        connectivity=1
    )

    shell_vals = filter_and_clip_peak_values(
        peak_3d[shell_only], peakref[0], peakref[1], low_margin, high_margin
    )
    bulk_vals = filter_and_clip_peak_values(
        peak_3d[bulk_only], peakref[0], peakref[1], low_margin, high_margin
    )

    shell_avg_ni = float(np.mean(peak_to_ni(shell_vals, ni_peak_lo, ni_peak_hi))) if shell_vals.size > 0 else np.nan
    bulk_avg_ni = float(np.mean(peak_to_ni(bulk_vals, ni_peak_lo, ni_peak_hi))) if bulk_vals.size > 0 else np.nan

    bulk_voxel_energy = analyze_voxel_peak_energy(
        bulk_only, peak_3d, peakref[0], peakref[1], low_margin, high_margin
    )
    shell_voxel_energy = analyze_voxel_peak_energy(
        shell_only, peak_3d, peakref[0], peakref[1], low_margin, high_margin
    )

    shell_h3_stat = analyze_domains_metric(shell_h3, peak_3d, voxel_um, peakref[0], peakref[1], low_margin, high_margin, ni_peak_lo, ni_peak_hi)
    shell_h2_stat = analyze_domains_metric(shell_h2, peak_3d, voxel_um, peakref[0], peakref[1], low_margin, high_margin, ni_peak_lo, ni_peak_hi)
    bulk_h3_stat = analyze_domains_metric(bulk_h3, peak_3d, voxel_um, peakref[0], peakref[1], low_margin, high_margin, ni_peak_lo, ni_peak_hi)
    bulk_h2_stat = analyze_domains_metric(bulk_h2, peak_3d, voxel_um, peakref[0], peakref[1], low_margin, high_margin, ni_peak_lo, ni_peak_hi)

    shell_h3_voxels = int(np.count_nonzero(shell_h3))
    shell_h2_voxels = int(np.count_nonzero(shell_h2))
    bulk_h3_voxels = int(np.count_nonzero(bulk_h3))
    bulk_h2_voxels = int(np.count_nonzero(bulk_h2))

    shell_phase_voxels = shell_h3_voxels + shell_h2_voxels
    bulk_phase_voxels = bulk_h3_voxels + bulk_h2_voxels

    shell_h3_fraction = float(shell_h3_voxels / shell_phase_voxels) if shell_phase_voxels > 0 else np.nan
    bulk_h3_fraction = float(bulk_h3_voxels / bulk_phase_voxels) if bulk_phase_voxels > 0 else np.nan

    return {
        "particle_folder_path": reg_dir,
        "particle_name": os.path.basename(os.path.dirname(reg_dir)) + f"_registered{os.path.basename(reg_dir).replace('registered', '')}",
        "effective_diameter_um": effective_diameter_um,
        "particle_mean_ni": particle_mean_ni,
        "particle_std_ni": particle_std_ni,
        "hist_bulk_voxel_energy": bulk_voxel_energy,
        "hist_shell_voxel_energy": shell_voxel_energy,
        "bulk_h3_width": bulk_h3_stat["width_um_list"],
        "shell_h3_width": shell_h3_stat["width_um_list"],
        "bulk_h2_width": bulk_h2_stat["width_um_list"],
        "shell_h2_width": shell_h2_stat["width_um_list"],
        "bulk_h3_volume": bulk_h3_stat["volume_um3_list"],
        "shell_h3_volume": shell_h3_stat["volume_um3_list"],
        "bulk_h2_volume": bulk_h2_stat["volume_um3_list"],
        "shell_h2_volume": shell_h2_stat["volume_um3_list"],
        "summary_shell_h3_count": shell_h3_stat["count"],
        "summary_bulk_h3_count": bulk_h3_stat["count"],
        "summary_shell_h2_count": shell_h2_stat["count"],
        "summary_bulk_h2_count": bulk_h2_stat["count"],
        "summary_total_h3_count": analyze_domains_metric(
            h3_mask_3d, peak_3d, voxel_um, peakref[0], peakref[1], low_margin, high_margin, ni_peak_lo, ni_peak_hi
        )["count"],
        "summary_total_h2_count": analyze_domains_metric(
            h2_mask_3d, peak_3d, voxel_um, peakref[0], peakref[1], low_margin, high_margin, ni_peak_lo, ni_peak_hi
        )["count"],
        "summary_shell_h3_fraction": shell_h3_fraction,
        "summary_bulk_h3_fraction": bulk_h3_fraction,
        "summary_shell_avg_ni": shell_avg_ni,
        "summary_bulk_avg_ni": bulk_avg_ni,
        "min_domain_voxels_after_intersection": int(min_domain_voxels_after_intersection),
        "h2_peaksite_values": h2_peaksite_values,
        "h3_peaksite_values": h3_peaksite_values,
        "h2_peaksite_mean": h2_peaksite_mean,
        "h3_peaksite_mean": h3_peaksite_mean,
        "h2_h3_split_energy": h2_h3_split_energy,
        "particle_soc": particle_soc,
        "xrd_h3_fraction": xrd_h3_fraction,
        "reg_dir": reg_dir,
        "roi_mask_dir": roi_mask_dir,
        "h3_dir": h3_dir,
        "h2_dir": h2_dir,
        "shell_dir": shell_dir,
        "bulk_dir": bulk_dir,
    }


def set_basic_style(ws):
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_list_tab(ws, particle_results, value_key):
    ws.cell(row=1, column=1, value="particle_folder_path").font = Font(bold=True)
    ws.cell(row=2, column=1, value="effective_diameter_um").font = Font(bold=True)
    ws.cell(row=3, column=1, value="particle_mean_ni").font = Font(bold=True)

    for idx, p in enumerate(particle_results, start=2):
        ws.cell(row=1, column=idx, value=p["particle_folder_path"])
        ws.cell(row=2, column=idx, value=p["effective_diameter_um"])
        ws.cell(row=3, column=idx, value=p["particle_mean_ni"])

        vals = p.get(value_key, [])
        if len(vals) == 0:
            ws.cell(row=4, column=idx, value="NO_DATA")
        else:
            for r, v in enumerate(vals, start=4):
                ws.cell(row=r, column=idx, value=v)

    set_basic_style(ws)


def write_phase_peaksite_tab(ws, particle_results, phase_name):
    phase_lower = phase_name.lower()

    values_key = f"{phase_lower}_peaksite_values"
    mean_key = f"{phase_lower}_peaksite_mean"

    ws.cell(row=1, column=1, value="입자경로").font = Font(bold=True)
    ws.cell(row=2, column=1, value="입자사이즈").font = Font(bold=True)
    ws.cell(row=3, column=1, value=f"{phase_name} peak_site 평균값").font = Font(bold=True)
    ws.cell(row=4, column=1, value=f"{phase_name}로 지정된 복셀들의 전체 peak_site").font = Font(bold=True)

    for idx, p in enumerate(particle_results, start=2):
        ws.cell(row=1, column=idx, value=p["particle_folder_path"])
        ws.cell(row=2, column=idx, value=p["effective_diameter_um"])
        ws.cell(row=3, column=idx, value=p.get(mean_key, np.nan))

        vals = p.get(values_key, [])
        if len(vals) == 0:
            ws.cell(row=4, column=idx, value="NO_DATA")
        else:
            for r, v in enumerate(vals, start=4):
                ws.cell(row=r, column=idx, value=v)

    set_basic_style(ws)


def write_summary_tab(ws, particle_results, analysis_mode):
    if analysis_mode == "pristine_fullcharge":
        headers = [
            "입자 폴더 경로",
            "입자실효지름",
            "입자 평균 니켈 산화수",
            "입자 평균 SOC",
            "입자 평균 SOC (%)",
            "입자 Ni 산화수 표준편차",
            "h2_h3 split energy",
        ]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=1, column=c, value=h).font = Font(bold=True)

        for r, p in enumerate(particle_results, start=2):
            particle_soc = p.get("particle_soc", np.nan)

            ws.cell(row=r, column=1, value=p["particle_folder_path"])
            ws.cell(row=r, column=2, value=p["effective_diameter_um"])
            ws.cell(row=r, column=3, value=p["particle_mean_ni"])
            ws.cell(row=r, column=4, value=particle_soc)
            ws.cell(row=r, column=5, value=particle_soc * 100 if np.isfinite(particle_soc) else np.nan)
            ws.cell(row=r, column=6, value=p["particle_std_ni"])
            ws.cell(row=r, column=7, value=p.get("h2_h3_split_energy", np.nan))

    else:
        headers = [
            "입자 폴더 경로",
            "입자실효지름",
            "입자 평균 니켈 산화수",
            "입자 평균 SOC",
            "입자 평균 SOC (%)",
            "입자 Ni 산화수 표준편차",
            "phase domain 최소 복셀 수",
            "shell avg Ni 산화수",
            "Bulk avg Ni 산화수",
            "shell H3 domain 갯수",
            "bulk H3 domain 갯수",
            "shell H2 domain 갯수",
            "bulk H2 domain 갯수",
            "전체 H3 domain 갯수",
            "전체 H2 domain 갯수",
            "shell에서의 H3 fraction [H3/(H2+H3)]",
            "bulk에서의 H3 fraction [H3/(H2+H3)]",
            "h2_h3 split energy",
        ]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=1, column=c, value=h).font = Font(bold=True)

        for r, p in enumerate(particle_results, start=2):
            particle_soc = p.get("particle_soc", np.nan)

            ws.cell(row=r, column=1, value=p["particle_folder_path"])
            ws.cell(row=r, column=2, value=p["effective_diameter_um"])
            ws.cell(row=r, column=3, value=p["particle_mean_ni"])
            ws.cell(row=r, column=4, value=particle_soc)
            ws.cell(row=r, column=5, value=particle_soc * 100 if np.isfinite(particle_soc) else np.nan)
            ws.cell(row=r, column=6, value=p["particle_std_ni"])
            ws.cell(row=r, column=7, value=p.get("min_domain_voxels_after_intersection", np.nan))
            ws.cell(row=r, column=8, value=p["summary_shell_avg_ni"])
            ws.cell(row=r, column=9, value=p["summary_bulk_avg_ni"])
            ws.cell(row=r, column=10, value=p["summary_shell_h3_count"])
            ws.cell(row=r, column=11, value=p["summary_bulk_h3_count"])
            ws.cell(row=r, column=12, value=p["summary_shell_h2_count"])
            ws.cell(row=r, column=13, value=p["summary_bulk_h2_count"])
            ws.cell(row=r, column=14, value=p["summary_total_h3_count"])
            ws.cell(row=r, column=15, value=p["summary_total_h2_count"])
            ws.cell(row=r, column=16, value=p["summary_shell_h3_fraction"])
            ws.cell(row=r, column=17, value=p["summary_bulk_h3_fraction"])
            ws.cell(row=r, column=18, value=p.get("h2_h3_split_energy", np.nan))

    set_basic_style(ws)

def save_requested_workbook(particle_results, out_xlsx, analysis_mode):
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    if analysis_mode == "phase_separation":
        sheet_order = [
            ("histogram bulk", "hist_bulk_voxel_energy"),
            ("histogram shell", "hist_shell_voxel_energy"),
            ("bulk H3 domain 폭", "bulk_h3_width"),
            ("shell H3 domain 폭", "shell_h3_width"),
            ("bulk H2 domain 폭", "bulk_h2_width"),
            ("shell H2 domain 폭", "shell_h2_width"),
            ("bulk H3 domain 부피", "bulk_h3_volume"),
            ("shell H3 domain 부피", "shell_h3_volume"),
            ("bulk H2 domain 부피", "bulk_h2_volume"),
            ("shell H2 domain 부피", "shell_h2_volume"),
        ]

        for sheet_name, key in sheet_order:
            ws = wb.create_sheet(title=sheet_name)
            write_list_tab(ws, particle_results, key)

        ws_h2 = wb.create_sheet(title="H2")
        write_phase_peaksite_tab(ws_h2, particle_results, "H2")

        ws_h3 = wb.create_sheet(title="H3")
        write_phase_peaksite_tab(ws_h3, particle_results, "H3")

    ws_sum = wb.create_sheet(title="summary")
    write_summary_tab(ws_sum, particle_results, analysis_mode)

    wb.save(out_xlsx)


def main():
    Tk().withdraw()

    analysis_mode = ask_analysis_mode()
    profiles = ask_bnl_profiles()
    phase_fraction_cfg = ask_phase_fraction_excel()
    peakref, low_margin, high_margin = ask_peak_base_and_margin()
    voxel_um = ask_voxel_size_um()
    min_domain_voxels_after_intersection = ask_min_domain_voxels_after_intersection()
    parents = ask_parent_directories_with_profiles(profiles)
    out_xlsx = ask_output_excel_path()

    if not parents:
        raise ValueError("❌ 분석할 상위 폴더가 없습니다.")

    particle_results = []

    for parent_item in parents:
        parent = parent_item["folder"]
        profile_name = parent_item["profile_name"]
        prof = profiles[profile_name]

        print(f"\n### 상위 폴더 처리 중: {parent} [profile={profile_name}] [mode={analysis_mode}]")
        print(f"### phase domain 최소 복셀 수 = {min_domain_voxels_after_intersection}")

        if phase_fraction_cfg is not None:
            print(f"### in situ XRD fraction Excel = {phase_fraction_cfg['path']}")
        else:
            print("### in situ XRD fraction Excel 미적용: h2_h3 split energy = NaN")

        star_folders = find_star_folders_recursively(parent)
        if not star_folders:
            print(f"[SKIP] ★ 폴더가 없습니다: {parent}")
            continue

        for cur_root, star_name in star_folders:
            n = extract_registered_number_from_star_folder(star_name)
            if n is None:
                continue

            reg_dir = os.path.join(cur_root, f"registered{n}")
            roi_mask_dir = os.path.join(reg_dir, f"{n}")

            if not os.path.isdir(reg_dir):
                print(f"[SKIP] registered 없음: {reg_dir}")
                continue
            if not os.path.isdir(roi_mask_dir):
                print(f"[SKIP] roi mask dir 없음: {roi_mask_dir}")
                continue

            try:
                result = process_registered_particle(
                    reg_dir=reg_dir,
                    roi_mask_dir=roi_mask_dir,
                    profiles_cfg=prof,
                    phase_fraction_cfg=phase_fraction_cfg,
                    peakref=peakref,
                    low_margin=low_margin,
                    high_margin=high_margin,
                    voxel_um=voxel_um,
                    analysis_mode=analysis_mode,
                    min_domain_voxels_after_intersection=min_domain_voxels_after_intersection,
                )
                particle_results.append(result)
                print(
                    f"[OK] {star_name} -> {result['particle_name']} | "
                    f"h2_h3 split energy = {result.get('h2_h3_split_energy', np.nan)}"
                )
            except Exception as e:
                print(f"[FAIL] {star_name}: {e}")

    if len(particle_results) == 0:
        raise ValueError("❌ 최종적으로 분석된 입자가 없습니다.")

    particle_results = sorted(
        particle_results,
        key=lambda x: (str(x["particle_folder_path"]), float(x["effective_diameter_um"]))
    )

    save_requested_workbook(particle_results, out_xlsx, analysis_mode)

    messagebox.showinfo("완료", f"엑셀 저장 완료:\n{out_xlsx}")


if __name__ == "__main__":
    main()