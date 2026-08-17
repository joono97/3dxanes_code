#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import glob
import numpy as np
import h5py
import warnings
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
import matplotlib.pyplot as plt
from skimage import io, color, img_as_ubyte
from skimage.measure import label, find_contours
from numpy.polynomial import polynomial as poly
from openpyxl import load_workbook


STATS_RANGE = (8340.0, 8360.0)
HIST_BINS = 2000

warnings.filterwarnings('ignore', message='.*is a low contrast image')


# ============================================================
# Basic utilities
# ============================================================

def find_registered_folders(parent_dir):
    items = []
    for name in os.listdir(parent_dir):
        full = os.path.join(parent_dir, name)
        if os.path.isdir(full):
            m = re.fullmatch(r"registered(\d+)", name)
            if m:
                items.append((int(m.group(1)), full))
    return sorted(items, key=lambda x: x[0])


def guess_roi_mask_dir(registered_dir, n):
    cand = os.path.join(registered_dir, f"{n}")
    if os.path.isdir(cand) and glob.glob(os.path.join(cand, "roi_mask_slice_*.npy")):
        return cand
    return None


def normalize_mask_to_shape(mask_2d, target_shape):
    th, tw = int(target_shape[0]), int(target_shape[1])
    out = np.zeros((th, tw), dtype=bool)
    mh, mw = mask_2d.shape
    hh, ww = min(th, mh), min(tw, mw)
    out[:hh, :ww] = mask_2d[:hh, :ww]
    return out


def normalize_array_to_shape(arr_2d, target_shape, fill_value=0):
    th, tw = int(target_shape[0]), int(target_shape[1])
    out = np.full((th, tw), fill_value, dtype=arr_2d.dtype)
    ah, aw = arr_2d.shape
    hh, ww = min(th, ah), min(tw, aw)
    out[:hh, :ww] = arr_2d[:hh, :ww]
    return out


def normalize_rgb_to_shape(arr_3d, target_shape):
    th, tw = int(target_shape[0]), int(target_shape[1])
    out = np.zeros((th, tw, 3), dtype=arr_3d.dtype)
    ah, aw = arr_3d.shape[:2]
    hh, ww = min(th, ah), min(tw, aw)
    out[:hh, :ww, :] = arr_3d[:hh, :ww, :]
    return out


def crop_3d_energy_stack_to_shape(arr_3d, target_shape):
    th, tw = int(target_shape[0]), int(target_shape[1])
    return arr_3d[:, :th, :tw]


# ============================================================
# GUI inputs
# ============================================================

def ask_parent_directories():
    parents = []
    while True:
        folder = filedialog.askdirectory(title="상위 폴더 선택 (취소 누르면 선택 종료)")
        if not folder:
            break
        if folder not in parents:
            parents.append(folder)
        more = messagebox.askyesno("폴더 추가 선택", "다른 상위 폴더를 추가로 선택하시겠습니까?")
        if not more:
            break
    return parents


def ask_stats_range(default_low=8340.0, default_high=8360.0):
    root = tk.Tk()
    root.withdraw()
    low = simpledialog.askfloat("STATS_RANGE LOW 입력", "stats range low", initialvalue=float(default_low))
    high = simpledialog.askfloat("STATS_RANGE HIGH 입력", "stats range high", initialvalue=float(default_high))
    low = float(default_low) if low is None else float(low)
    high = float(default_high) if high is None else float(high)
    if low > high:
        low, high = high, low
    return (low, high)


def ask_margins(default_low=0.0, default_high=0.0):
    root = tk.Tk()
    root.withdraw()
    lm = simpledialog.askfloat("LOW_MARGIN 입력", "low_margin", initialvalue=float(default_low))
    hm = simpledialog.askfloat("HIGH_MARGIN 입력", "high_margin", initialvalue=float(default_high))
    low_margin = float(default_low) if lm is None else float(lm)
    high_margin = float(default_high) if hm is None else float(hm)
    return low_margin, high_margin


def ask_energy_axis_options():
    root = tk.Tk()
    root.withdraw()
    reverse_stack = messagebox.askyesno("Image stack reverse?", "TIFF stack을 [::-1,:,:]로 뒤집을까요?")
    reverse_energy = messagebox.askyesno("Energy file reverse?", "energy.txt 순서를 역순으로 뒤집을까요?")
    return reverse_stack, reverse_energy


def ask_fit_parameters(default_ev_step=0.5, default_peakref_low=8352.1, default_peakref_high=8355.3):
    root = tk.Tk()
    root.withdraw()

    ev_step = simpledialog.askfloat(
        "ev_step 입력",
        "피크 피팅에 사용할 에너지 스텝(eV)을 입력하세요.",
        initialvalue=float(default_ev_step)
    )
    if ev_step is None:
        ev_step = float(default_ev_step)

    peakref_low = simpledialog.askfloat(
        "peakref low 입력",
        "peakref의 low 값을 입력하세요.",
        initialvalue=float(default_peakref_low)
    )
    if peakref_low is None:
        peakref_low = float(default_peakref_low)

    peakref_high = simpledialog.askfloat(
        "peakref high 입력",
        "peakref의 high 값을 입력하세요.",
        initialvalue=float(default_peakref_high)
    )
    if peakref_high is None:
        peakref_high = float(default_peakref_high)

    if peakref_low > peakref_high:
        peakref_low, peakref_high = peakref_high, peakref_low

    return float(ev_step), (float(peakref_low), float(peakref_high))


def ask_min_connected_voxels_after_combine(default_value=3):
    root = tk.Tk()
    root.withdraw()
    s = simpledialog.askstring(
        "도메인 최소 연결 복셀 수",
        "H2/H3 phase domain을 shell/bulk overlap이 더 많은 쪽으로 배정할 때,\n"
        "유지할 최소 연결 복셀 수를 입력하세요.\n"
        "예: 3 이면 H2/H3 원래 domain 중 1~2 voxel 도메인은 제외",
        initialvalue=str(int(default_value))
    )
    try:
        v = int(float(s))
    except Exception:
        v = int(default_value)
    if v < 1:
        v = 1
    return v


def ask_soc_calibration_peak_sites(default_soc0=8353.5, default_soc100=8355.2):
    root = tk.Tk()
    root.withdraw()

    soc0_peak = simpledialog.askfloat(
        "SOC0 peak_site 입력",
        "SOC0에 해당하는 peak_site(eV)를 입력하세요.",
        initialvalue=float(default_soc0)
    )
    soc100_peak = simpledialog.askfloat(
        "SOC100 peak_site 입력",
        "SOC100에 해당하는 peak_site(eV)를 입력하세요.",
        initialvalue=float(default_soc100)
    )

    soc0_peak = float(default_soc0) if soc0_peak is None else float(soc0_peak)
    soc100_peak = float(default_soc100) if soc100_peak is None else float(soc100_peak)

    if soc0_peak == soc100_peak:
        raise ValueError("SOC0 peak_site와 SOC100 peak_site가 같으면 SOC 계산을 할 수 없습니다.")

    return soc0_peak, soc100_peak


def ask_phase_fraction_excel_path():
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(
        title="SOC-H2/H3 상분율 Excel 파일 선택",
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
    )
    if not path:
        raise ValueError("SOC-H2/H3 상분율 Excel 파일을 선택하지 않았습니다.")
    return path


# ============================================================
# Mask/domain utilities
# ============================================================

def remove_small_components_3d(mask_3d, min_connected_voxels=3, connectivity=1):
    mask_3d = np.asarray(mask_3d, dtype=bool)

    if mask_3d.ndim != 3 or not np.any(mask_3d):
        return mask_3d.copy()

    min_connected_voxels = max(1, int(min_connected_voxels))
    if min_connected_voxels <= 1:
        return mask_3d.copy()

    lbl = label(mask_3d.astype(np.uint8), connectivity=connectivity)
    if lbl.max() == 0:
        return mask_3d.copy()

    counts = np.bincount(lbl.ravel())
    keep_labels = np.where(counts >= min_connected_voxels)[0]
    keep_labels = keep_labels[keep_labels != 0]

    if keep_labels.size == 0:
        return np.zeros_like(mask_3d, dtype=bool)

    return np.isin(lbl, keep_labels)


def assign_phase_domains_to_shell_or_bulk(phase_mask_3d, shell_mask_3d, bulk_mask_3d,
                                          min_connected_voxels=1, connectivity=1):
    phase_mask_3d = np.asarray(phase_mask_3d, dtype=bool)
    shell_mask_3d = np.asarray(shell_mask_3d, dtype=bool)
    bulk_mask_3d = np.asarray(bulk_mask_3d, dtype=bool)

    shell_assigned = np.zeros_like(phase_mask_3d, dtype=bool)
    bulk_assigned = np.zeros_like(phase_mask_3d, dtype=bool)

    if phase_mask_3d.ndim != 3 or not np.any(phase_mask_3d):
        return shell_assigned, bulk_assigned

    min_connected_voxels = max(1, int(min_connected_voxels))

    lbl = label(phase_mask_3d.astype(np.uint8), connectivity=connectivity)
    if lbl.max() == 0:
        return shell_assigned, bulk_assigned

    for lab in range(1, int(lbl.max()) + 1):
        dom = (lbl == lab)
        dom_voxels = int(np.count_nonzero(dom))

        if dom_voxels < min_connected_voxels:
            continue

        shell_count = int(np.count_nonzero(dom & shell_mask_3d))
        bulk_count = int(np.count_nonzero(dom & bulk_mask_3d))

        if shell_count == 0 and bulk_count == 0:
            continue

        if bulk_count > shell_count:
            bulk_assigned[dom] = True
        else:
            shell_assigned[dom] = True

    return shell_assigned, bulk_assigned


def choose_mask_folder(roi_mask_dir, kind):
    cands = []
    for name in os.listdir(roi_mask_dir):
        full = os.path.join(roi_mask_dir, name)
        if not os.path.isdir(full):
            continue
        if not glob.glob(os.path.join(full, "filtered_mask_slice_*.npy")):
            continue
        lowname = name.lower()
        matched = False
        if kind == "upper":
            matched = ("00" in lowname)
        elif kind == "lower":
            matched = ("100" in lowname)
        elif kind == "shell":
            matched = ("shell" in lowname)
        elif kind == "bulk":
            matched = ("bulk" in lowname)
        if matched:
            cands.append(full)
    cands.sort()
    return cands[0] if cands else None


def load_mask_stack(mask_dir, slice_indices, target_shapes):
    masks = []
    for sidx, target_shape in zip(slice_indices, target_shapes):
        path = os.path.join(mask_dir, f"filtered_mask_slice_{sidx:04d}.npy")
        if os.path.exists(path):
            m = np.load(path).astype(bool)
            if m.shape != target_shape:
                m = normalize_mask_to_shape(m, target_shape)
        else:
            m = np.zeros(target_shape, dtype=bool)
        masks.append(m)
    return np.stack(masks, axis=0).astype(bool)


# ============================================================
# Energy / SOC / H2-H3 split energy
# ============================================================

def align_energy_axis(img_stack, eng, reverse_stack=False, reverse_energy=False, verbose=True):
    eng = np.asarray(eng, dtype=float).ravel()
    if reverse_stack:
        img_stack = img_stack[::-1, :, :]
    if reverse_energy:
        eng = eng[::-1]
    if eng[0] > eng[-1]:
        eng = eng[::-1]
        img_stack = img_stack[::-1, :, :]
    if verbose:
        print(f"[i] aligned energy range: {eng[0]:.4f} -> {eng[-1]:.4f} eV, N={len(eng)}")
    return img_stack, eng


def compute_stats_single_criterion(roi_values, low, high, low_margin, high_margin):
    roi_values = np.asarray(roi_values, dtype=float)
    roi_values = roi_values[np.isfinite(roi_values)]
    eff_low = float(low) - float(low_margin)
    eff_high = float(high) + float(high_margin)
    valid_mask = (roi_values >= eff_low) & (roi_values <= eff_high)
    if not np.any(valid_mask):
        return np.array([], dtype=float)
    v = roi_values[valid_mask].astype(float, copy=False)
    v = v.copy()
    v[v < low] = low
    v[v > high] = high
    return v


def filter_energy_in_range(values, low, high):
    values = np.asarray(values, dtype=float)
    values = values[np.isfinite(values)]
    return values[(values >= float(low)) & (values <= float(high))]


def load_soc_h3_fraction_table(excel_path):
    wb = load_workbook(excel_path, data_only=True, read_only=True)

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_row_idx = None
        headers = None

        for i, row in enumerate(rows[:15]):
            header_candidates = [str(x).strip().lower() if x is not None else "" for x in row]
            if any("soc" in h for h in header_candidates) and any("h3" in h for h in header_candidates):
                header_row_idx = i
                headers = header_candidates
                break

        if header_row_idx is None:
            continue

        soc_col = None
        h2_col = None
        h3_col = None
        h3_frac_col = None

        for j, h in enumerate(headers):
            if "soc" in h and soc_col is None:
                soc_col = j
            if "h2" in h and h2_col is None:
                h2_col = j
            if "h3" in h and h3_col is None:
                h3_col = j
            if "h3" in h and any(k in h for k in ["fraction", "frac", "ratio", "비율", "%"]):
                h3_frac_col = j

        if soc_col is None:
            continue

        soc_list = []
        h3_frac_list = []

        for row in rows[header_row_idx + 1:]:
            if row is None or soc_col >= len(row):
                continue

            try:
                soc = float(row[soc_col])
            except Exception:
                continue

            h3_frac = None

            if h3_frac_col is not None and h3_frac_col < len(row):
                try:
                    h3_frac = float(row[h3_frac_col])
                except Exception:
                    h3_frac = None

            if h3_frac is None and h2_col is not None and h3_col is not None:
                if h2_col < len(row) and h3_col < len(row):
                    try:
                        h2v = float(row[h2_col])
                        h3v = float(row[h3_col])
                        denom = h2v + h3v
                        if denom > 0:
                            h3_frac = h3v / denom
                    except Exception:
                        h3_frac = None

            if h3_frac is None or not np.isfinite(h3_frac):
                continue

            if h3_frac > 1.0:
                h3_frac = h3_frac / 100.0

            soc_list.append(float(soc))
            h3_frac_list.append(float(np.clip(h3_frac, 0.0, 1.0)))

        if len(soc_list) >= 2:
            soc_arr = np.asarray(soc_list, dtype=float)
            h3_arr = np.asarray(h3_frac_list, dtype=float)
            order = np.argsort(soc_arr)
            return soc_arr[order], h3_arr[order]

    raise ValueError("Excel에서 SOC-H3 fraction table을 찾지 못했습니다. SOC 열과 H3 fraction 또는 H2/H3 열이 필요합니다.")


def peak_to_soc_percent(mean_peak, soc0_peak_site, soc100_peak_site):
    if not np.isfinite(mean_peak):
        return np.nan

    denom = float(soc100_peak_site) - float(soc0_peak_site)
    if denom == 0:
        return np.nan

    soc = (float(mean_peak) - float(soc0_peak_site)) / denom * 100.0
    return float(np.clip(soc, 0.0, 100.0))


def interpolate_h3_fraction_from_soc(particle_soc, soc_arr, h3_frac_arr):
    if not np.isfinite(particle_soc):
        return np.nan
    return float(np.interp(float(particle_soc), soc_arr, h3_frac_arr))


def get_separation_energy_from_h3_fraction(peak_values, target_h3_fraction, stats_range):
    """
    H2/H3 split energy 계산.
    H3가 높은 peak_site 쪽이라고 가정하고,
    peak_site >= split_energy 인 voxel fraction이 target_h3_fraction이 되도록 한다.
    """
    values = np.asarray(peak_values, dtype=float)
    values = values[np.isfinite(values)]

    low, high = float(stats_range[0]), float(stats_range[1])
    values = values[(values >= low) & (values <= high)]

    if values.size == 0 or not np.isfinite(target_h3_fraction):
        return np.nan

    q = 1.0 - float(np.clip(target_h3_fraction, 0.0, 1.0))
    return float(np.quantile(values, q))


def get_particle_mean_peaksite_in_stats_range(peak_raw_3d, binary_mask_list, stats_range):
    vals_all = []
    low, high = float(stats_range[0]), float(stats_range[1])

    for raw2d, roi2d in zip(peak_raw_3d, binary_mask_list):
        vals = np.asarray(raw2d[roi2d], dtype=float)
        vals = vals[np.isfinite(vals)]
        vals = vals[(vals >= low) & (vals <= high)]
        if vals.size:
            vals_all.append(vals)

    if not vals_all:
        return np.nan

    allv = np.concatenate(vals_all)
    return float(np.mean(allv))


def collect_roi_peak_values(peak_raw_3d, binary_mask_list, stats_range):
    vals_all = []
    low, high = float(stats_range[0]), float(stats_range[1])

    for raw2d, roi2d in zip(peak_raw_3d, binary_mask_list):
        vals = np.asarray(raw2d[roi2d], dtype=float)
        vals = vals[np.isfinite(vals)]
        vals = vals[(vals >= low) & (vals <= high)]
        if vals.size:
            vals_all.append(vals)

    if not vals_all:
        return np.array([], dtype=float)

    return np.concatenate(vals_all).astype(float)


# ============================================================
# Image making / saving
# ============================================================

def label_to_rgb_slice(label2d, top_label_color_map=None):
    if top_label_color_map is None:
        top_label_color_map = {}

    h, w = label2d.shape
    rgb = np.zeros((h, w, 3), dtype=np.float32)
    uniq = np.unique(label2d)
    uniq = uniq[uniq > 0]

    for lab in uniq:
        lab_int = int(lab)

        if lab_int in top_label_color_map:
            c = np.array(top_label_color_map[lab_int], dtype=np.float32)
        else:
            rng = np.random.default_rng(lab_int)
            c = rng.uniform(0.15, 1.0, size=3).astype(np.float32)
            if c.mean() < 0.35:
                c = np.clip(c + 0.25, 0, 1)

        rgb[label2d == lab] = c

    return rgb


def make_core_vs_shell_rgb(shell2d, bulk2d, phase2d):
    rgb = np.zeros((phase2d.shape[0], phase2d.shape[1], 3), dtype=np.float32)
    rgb[shell2d & phase2d] = np.array([1.0, 1.0, 0.0], dtype=np.float32)
    rgb[bulk2d & phase2d] = np.array([1.0, 0.0, 0.0], dtype=np.float32)
    return rgb


def make_phase_field_rgb(h2_2d, h3_2d):
    rgb = np.zeros((h2_2d.shape[0], h2_2d.shape[1], 3), dtype=np.float32)
    rgb[h2_2d] = np.array([0.53, 0.81, 0.98], dtype=np.float32)
    rgb[h3_2d] = np.array([0.0, 0.0, 1.0], dtype=np.float32)
    return rgb


def make_single_color_mask_rgb(mask2d, color_rgb):
    rgb = np.zeros((mask2d.shape[0], mask2d.shape[1], 3), dtype=np.float32)
    rgb[mask2d] = np.array(color_rgb, dtype=np.float32)
    return rgb


def make_diverging_centered_rgb(raw2d, roi2d, center_x, stats_range, half_width=0.3):
    rgb = np.zeros((raw2d.shape[0], raw2d.shape[1], 3), dtype=np.float32)

    if not np.isfinite(center_x):
        return rgb

    low_stats, high_stats = float(stats_range[0]), float(stats_range[1])
    vals = np.asarray(raw2d, dtype=float)

    valid = roi2d & np.isfinite(vals) & (vals >= low_stats) & (vals <= high_stats)
    if not np.any(valid):
        return rgb

    cmap_low = float(center_x - half_width)
    cmap_high = float(center_x + half_width)

    clipped = vals.copy()
    clipped[valid] = np.clip(clipped[valid], cmap_low, cmap_high)

    orange = np.array([1.0, 165 / 255, 0.0], dtype=np.float32)
    white = np.array([1.0, 1.0, 1.0], dtype=np.float32)
    purple = np.array([128 / 255, 0.0, 128 / 255], dtype=np.float32)

    below = valid & (clipped <= center_x)
    above = valid & (clipped >= center_x)

    if half_width > 0:
        t_below = (clipped[below] - cmap_low) / half_width
        t_below = np.clip(t_below, 0.0, 1.0)

        t_above = (clipped[above] - center_x) / half_width
        t_above = np.clip(t_above, 0.0, 1.0)
    else:
        t_below = np.ones(np.count_nonzero(below), dtype=float)
        t_above = np.zeros(np.count_nonzero(above), dtype=float)

    if np.any(below):
        rgb[below] = (1.0 - t_below)[:, None] * orange + t_below[:, None] * white

    if np.any(above):
        rgb[above] = (1.0 - t_above)[:, None] * white + t_above[:, None] * purple

    return rgb


def get_diverging_side_masks(raw2d, roi2d, center_x, stats_range, half_width=0.3):
    vals = np.asarray(raw2d, dtype=float)
    low_stats, high_stats = float(stats_range[0]), float(stats_range[1])

    valid = roi2d & np.isfinite(vals) & (vals >= low_stats) & (vals <= high_stats)
    if not np.any(valid):
        z = np.zeros_like(roi2d, dtype=bool)
        return z, z

    cmap_low = float(center_x - half_width)
    cmap_high = float(center_x + half_width)

    clipped = vals.copy()
    clipped[valid] = np.clip(clipped[valid], cmap_low, cmap_high)

    orange_mask = valid & (clipped <= center_x)
    purple_mask = valid & (clipped >= center_x)

    return orange_mask, purple_mask


def get_boundary_pixels_4conn(mask2d):
    m = np.asarray(mask2d, dtype=bool)
    if not np.any(m):
        return np.zeros_like(m, dtype=bool)

    up = np.zeros_like(m, dtype=bool)
    down = np.zeros_like(m, dtype=bool)
    left = np.zeros_like(m, dtype=bool)
    right = np.zeros_like(m, dtype=bool)

    up[1:, :] = m[:-1, :]
    down[:-1, :] = m[1:, :]
    left[:, 1:] = m[:, :-1]
    right[:, :-1] = m[:, 1:]

    interior = m & up & down & left & right
    boundary = m & (~interior)
    return boundary


def ordered_boundary_centers_from_domain(domain_mask):
    domain_mask = np.asarray(domain_mask, dtype=bool)
    if not np.any(domain_mask):
        return np.empty((0, 2), dtype=float)

    boundary_mask = get_boundary_pixels_4conn(domain_mask)
    boundary_pts = np.argwhere(boundary_mask)
    if boundary_pts.shape[0] == 0:
        return np.empty((0, 2), dtype=float)

    contours = find_contours(domain_mask.astype(float), level=0.5)
    if len(contours) == 0:
        return boundary_pts.astype(float)

    contour = max(contours, key=lambda x: len(x))

    diff = contour[:, None, :] - boundary_pts[None, :, :]
    dist2 = np.sum(diff * diff, axis=2)
    nearest_idx = np.argmin(dist2, axis=1)
    ordered = boundary_pts[nearest_idx]

    keep = np.ones(len(ordered), dtype=bool)
    if len(ordered) > 1:
        keep[1:] = np.any(ordered[1:] != ordered[:-1], axis=1)
    ordered = ordered[keep]

    return ordered.astype(float)


def overlay_diverging_domain_centerlines(div_rgb, orange_mask_2d, purple_mask_2d, linewidth=0.3):
    rgb = np.asarray(div_rgb, dtype=float)
    h, w = rgb.shape[:2]
    dpi = 100.0

    fig = plt.figure(figsize=(w / dpi, h / dpi), dpi=dpi, frameon=False)
    ax = fig.add_axes([0, 0, 1, 1])
    ax.imshow(np.clip(rgb, 0, 1), interpolation='nearest')

    for phase_mask in [orange_mask_2d, purple_mask_2d]:
        if phase_mask is None or not np.any(phase_mask):
            continue

        lbl = label(phase_mask.astype(np.uint8), connectivity=1)
        uniq = np.unique(lbl)
        uniq = uniq[uniq > 0]

        for lab in uniq:
            domain_mask = (lbl == lab)
            pts = ordered_boundary_centers_from_domain(domain_mask)
            if pts.shape[0] < 2:
                continue

            ax.plot(pts[:, 1], pts[:, 0], color='black', linewidth=linewidth)

            if pts.shape[0] >= 3:
                ax.plot([pts[-1, 1], pts[0, 1]],
                        [pts[-1, 0], pts[0, 0]],
                        color='black', linewidth=linewidth)

    ax.set_axis_off()
    ax.set_xlim(-0.5, w - 0.5)
    ax.set_ylim(h - 0.5, -0.5)

    fig.canvas.draw()
    buf = np.frombuffer(fig.canvas.buffer_rgba(), dtype=np.uint8)
    arr = buf.reshape(fig.canvas.get_width_height()[1], fig.canvas.get_width_height()[0], 4)
    out = arr[:, :, :3].astype(np.float32) / 255.0
    plt.close(fig)
    return out


def save_rgb_stack(out_dir, slice_names, rgb_slices):
    os.makedirs(out_dir, exist_ok=True)
    for fn, rgb in zip(slice_names, rgb_slices):
        io.imsave(os.path.join(out_dir, f"{fn}.png"), img_as_ubyte(np.clip(rgb, 0, 1)))


def save_domain_sep_stack(out_dir, slice_names, mask_3d):
    os.makedirs(out_dir, exist_ok=True)

    lbl3d = label(mask_3d.astype(np.uint8), connectivity=1)

    counts = np.bincount(lbl3d.ravel())
    if counts.size > 0:
        counts[0] = 0

    labels = np.arange(len(counts))
    valid_labels = labels[counts > 0]
    sorted_labels = valid_labels[np.argsort(counts[valid_labels])[::-1]]

    top_colors = [
        np.array([0.53, 0.81, 0.98], dtype=np.float32),
        np.array([0.8, 0.6, 1.0], dtype=np.float32),
        np.array([0.6, 1.0, 0.6], dtype=np.float32),
    ]

    top_label_color_map = {}
    for rank, lab in enumerate(sorted_labels[:3]):
        top_label_color_map[int(lab)] = top_colors[rank]

    if len(sorted_labels) > 0:
        msg_parts = []
        for rank, lab in enumerate(sorted_labels[:3], start=1):
            msg_parts.append(f"{rank}등 label={int(lab)}, voxels={int(counts[lab])}")
        print(f"   domain color ranking for {out_dir}: " + " / ".join(msg_parts))

    for fn, lab2d in zip(slice_names, lbl3d):
        rgb = label_to_rgb_slice(lab2d, top_label_color_map=top_label_color_map)
        io.imsave(os.path.join(out_dir, f"{fn}.png"), img_as_ubyte(np.clip(rgb, 0, 1)))


def save_histogram_csv(hist_dir, values, stats_range, bins=HIST_BINS):
    os.makedirs(hist_dir, exist_ok=True)
    low, high = float(stats_range[0]), float(stats_range[1])
    values = np.asarray(values, dtype=float)
    values = values[np.isfinite(values)]
    hist, bins_arr = np.histogram(values, bins=bins, range=[low, high])
    center = (bins_arr[:-1] + bins_arr[1:]) / 2.0
    out_csv = os.path.join(hist_dir, 'registered_volume_histogram.csv')
    np.savetxt(out_csv, np.column_stack([center, hist]), delimiter=',', header='bin_center_eV,count', comments='')
    return out_csv


def save_histogram_raw_csv(hist_dir, values, filename='registered_volume_histogram_raw.csv'):
    os.makedirs(hist_dir, exist_ok=True)
    values = np.asarray(values, dtype=float)
    values = values[np.isfinite(values)]
    out_csv = os.path.join(hist_dir, filename)
    if values.size == 0:
        np.savetxt(out_csv, np.empty((0, 1)), delimiter=',', header='peak_site_eV_raw', comments='')
    else:
        np.savetxt(out_csv, values.reshape(-1, 1), delimiter=',', header='peak_site_eV_raw', comments='')
    return out_csv


def save_peaksite_energy_raw_csv(hist_dir, values, filename='peak_site_energy_raw.csv'):
    os.makedirs(hist_dir, exist_ok=True)
    values = np.asarray(values, dtype=float)
    values = values[np.isfinite(values)]
    out_csv = os.path.join(hist_dir, filename)
    if values.size == 0:
        np.savetxt(out_csv, np.empty((0, 1)), delimiter=',', header='peak_site_eV', comments='')
    else:
        np.savetxt(out_csv, values.reshape(-1, 1), delimiter=',', header='peak_site_eV', comments='')
    return out_csv


def save_probability_histogram_plot_01bin(hist_dir, peak_values, stats_range, bin_width=0.1):
    os.makedirs(hist_dir, exist_ok=True)

    values = np.asarray(peak_values, dtype=float)
    values = values[np.isfinite(values)]

    low, high = float(stats_range[0]), float(stats_range[1])
    values = values[(values >= low) & (values <= high)]

    if values.size == 0:
        print(f"⚠️ 0.1 bin histogram skipped: no valid values in {hist_dir}")
        return

    bins = np.arange(low, high + bin_width, bin_width)
    if bins[-1] < high:
        bins = np.append(bins, high)

    counts, bin_edges = np.histogram(values, bins=bins)
    total = np.sum(counts)

    if total <= 0:
        print(f"⚠️ 0.1 bin histogram skipped: zero counts in {hist_dir}")
        return

    prob = counts.astype(float) / float(total)
    centers = (bin_edges[:-1] + bin_edges[1:]) / 2.0

    out_csv = os.path.join(hist_dir, "peaksite_histogram_probability_bin0p1.csv")
    np.savetxt(
        out_csv,
        np.column_stack([centers, prob, counts]),
        delimiter=",",
        header="bin_center_eV,probability,count",
        comments=""
    )

    fig, ax = plt.subplots(figsize=(6, 4))
    ax.bar(centers, prob, width=bin_width, align="center", edgecolor="black", linewidth=0.3)
    ax.set_xlabel("peak_site (eV)")
    ax.set_ylabel("Probability")
    ax.set_title("Peak-site Probability Distribution, bin = 0.1 eV")
    ax.set_xlim(low, high)
    ax.set_ylim(bottom=0)
    fig.tight_layout()

    out_png = os.path.join(hist_dir, "peaksite_histogram_probability_bin0p1.png")
    fig.savefig(out_png, dpi=300)
    plt.close(fig)

    print(f"✅ 0.1 bin histogram probability plot saved: {out_png}")
    print(f"   - probability sum = {np.sum(prob):.6f}")


def save_h5_stack(h5_dir, slice_names, img_aligned_list, peaksite_list,
                  peaksite_raw_list, thickness_list, rgb_list, binary_mask_list, energy_list):
    os.makedirs(h5_dir, exist_ok=True)
    for fn, img_aligned, peaksite, peaksite_raw, thickness, rgb, binary_mask, energy in zip(
        slice_names, img_aligned_list, peaksite_list, peaksite_raw_list,
        thickness_list, rgb_list, binary_mask_list, energy_list
    ):
        out_path = os.path.join(h5_dir, f"image_dataset_id_{fn}_pos_00.h5")
        with h5py.File(out_path, 'w') as hf:
            hf.create_dataset('Img_aligned', data=img_aligned)
            hf.create_dataset('Peaksite', data=peaksite)
            hf.create_dataset('Peaksite_raw', data=peaksite_raw)
            hf.create_dataset('Thickness', data=thickness)
            hf.create_dataset('RGB_image', data=rgb)
            hf.create_dataset('BinaryMask', data=binary_mask.astype(np.uint8))
            hf.create_dataset('Energy', data=energy)


# ============================================================
# XANES fitting class
# ============================================================

class fit_xanes:
    def __init__(self, img, eng, peakref, color_flag):
        self.img = img
        self.eng = np.array(eng).ravel()
        self.peakref = list(peakref)
        self.pre_thick = np.zeros(self.img[0].shape)
        self.post_thick = np.zeros(self.img[0].shape)
        self.binary_mask = np.zeros(self.img[0].shape, dtype=bool)
        self.hsvconstant = np.array((2 / 3, 1, 0), dtype=float) if color_flag == 'rgb' else np.array((1 / 3, 1, 0), dtype=float)
        self.peaksite_raw = None
        self.peaksite = None
        self.rgb = None

    def set_thickness_pnts(self, eng_len):
        if eng_len == 101:
            self.pre_no = 30
            self.post_no = 91
        elif eng_len == 63:
            self.pre_no = 8
            self.post_no = 54
        elif eng_len == 21:
            self.pre_no = 2
            self.post_no = 18
        else:
            self.pre_no = max(1, eng_len // 5)
            self.post_no = max(self.pre_no + 1, eng_len - 5)

    def set_thickness(self):
        for i in range(self.pre_no):
            self.pre_thick += self.img[i]
        for i in range(self.post_no, len(self.img)):
            self.post_thick += self.img[i]
        self.post_mean = self.post_thick / max(1, (len(self.img) - self.post_no))
        self.pre_mean = self.pre_thick / max(1, self.pre_no)
        self.thickness = self.post_mean - self.pre_mean

    def polynomial_second_fit_separate(self, fit_num, ev_step):
        y = self.img.reshape(len(self.img), -1)
        y_new = np.zeros((fit_num * 2 + 1, y.shape[1]))
        mp_list = []
        for i in range(y.shape[1]):
            a = np.argmax(y[:, i])
            if a + fit_num + 1 > len(y):
                y_new[:, i] = y[len(y) - 2 * fit_num - 1:len(y) + 1, i]
            elif a - fit_num < 0:
                y_new[:, i] = y[0: 2 * fit_num + 1, i]
            else:
                y_new[:, i] = y[a - fit_num:a + fit_num + 1, i]
            mp_list.append(a)
        x = np.linspace(0, ev_step * (len(y_new) - 1), len(y_new))
        coefs, _ = poly.polyfit(x, y_new, 2, full=True)
        ind = [a - fit_num for a in mp_list]
        with np.errstate(divide='ignore', invalid='ignore'):
            x0 = -(coefs[1] / 2 / coefs[2]) + self.eng[ind]
        peaksite = x0.reshape(self.img.shape[1], self.img.shape[2])
        self.peaksite_raw = peaksite.astype(np.float32, copy=False)
        self.peaksite_raw[self.binary_mask == 0] = np.nan
        peaksite = np.clip(peaksite, float(self.peakref[0]), float(self.peakref[1]))
        peaksite[self.binary_mask == 0] = 0
        self.peaksite = peaksite

    def hsvcolormap(self, value=3):
        hsv = np.zeros((self.img.shape[1], self.img.shape[2], 3), dtype='float32')

        denom = float(self.peakref[1] - self.peakref[0])
        if denom == 0:
            denom = 1e-12

        conc = (self.peaksite - self.peakref[0]) / denom
        conc = np.nan_to_num(conc, nan=0.0, posinf=1.0, neginf=0.0)
        conc = np.clip(conc, 0.0, 1.0)

        thick = np.nan_to_num(self.thickness, nan=0.0, posinf=0.0, neginf=0.0)
        thick = np.clip(thick, 0.0, None)

        if value != 0:
            v = thick * value
        else:
            max_thick = np.nanmax(thick) if np.any(np.isfinite(thick)) else 1.0
            if max_thick <= 0:
                max_thick = 1.0
            v = thick / max_thick

        v = np.nan_to_num(v, nan=0.0, posinf=1.0, neginf=0.0)
        v = np.clip(v, 0.0, 1.0)

        hsv[:, :, 0] = conc * self.hsvconstant[0]
        hsv[:, :, 1] = 1.0
        hsv[:, :, 2] = v

        self.rgb = color.hsv2rgb(hsv)
        self.rgb = np.nan_to_num(self.rgb, nan=0.0, posinf=1.0, neginf=0.0)
        self.rgb = np.clip(self.rgb, 0.0, 1.0)


# ============================================================
# Main particle processing
# ============================================================

def process_registered_with_saved_masks(parent_dir, reg_dir, n, energy_path, peakref, fit_num,
                                        color_flag, sli_range, ev_step,
                                        stats_range, low_margin, high_margin,
                                        reverse_stack, reverse_energy,
                                        min_connected_voxels_after_combine=3,
                                        soc0_peak_site=None,
                                        soc100_peak_site=None,
                                        phase_soc_arr=None,
                                        phase_h3_frac_arr=None):
    roi_mask_dir = guess_roi_mask_dir(reg_dir, n)

    use_full_roi = roi_mask_dir is None
    if use_full_roi:
        print(f"⚠️ registered{n}: ROI mask 폴더가 없습니다.")
        print("   -> 전체 이미지 영역을 ROI로 간주하고 3D_x_fit_3★ 전체 컬러맵/histogram/h5만 생성합니다.")
        upper_dir = None
        lower_dir = None
        shell_dir = None
        bulk_dir = None
    else:
        upper_dir = choose_mask_folder(roi_mask_dir, 'upper')
        lower_dir = choose_mask_folder(roi_mask_dir, 'lower')
        shell_dir = choose_mask_folder(roi_mask_dir, 'shell')
        bulk_dir = choose_mask_folder(roi_mask_dir, 'bulk')

    missing_masks = []
    if upper_dir is None:
        missing_masks.append('upper')
    if lower_dir is None:
        missing_masks.append('lower')
    if shell_dir is None:
        missing_masks.append('shell')
    if bulk_dir is None:
        missing_masks.append('bulk')

    if missing_masks:
        print(f"⚠️ registered{n}: missing mask folders -> {', '.join(missing_masks)}")
        print("   -> ROI 마스크 기반 전체 폴더(3D_x_fit_3★)만 생성하고 상/하단/쉘/벌크 개별 폴더는 건너뜁니다.")

    tiff_files = sorted(glob.glob(os.path.join(reg_dir, '*.tif')) + glob.glob(os.path.join(reg_dir, '*.tiff')))
    if not tiff_files:
        raise ValueError(f"No TIFF files found in {reg_dir}")

    eng = np.loadtxt(energy_path)
    eng = np.asarray(eng, dtype=float).ravel()

    imgs_rgb = []
    peak_raw_list = []
    peaksite_list = []
    thickness_list = []
    binary_mask_list = []
    img_aligned_list = []
    energy_list = []
    slice_names = []
    slice_indices = []
    target_shapes = []

    for idx, slice_path in enumerate(tiff_files, start=1):
        img_raw = io.imread(slice_path)
        if sli_range is not None:
            img_raw = img_raw[sli_range[0]:sli_range[1], :, :]
        img, eng_aligned = align_energy_axis(
            img_raw, eng,
            reverse_stack=reverse_stack,
            reverse_energy=reverse_energy,
            verbose=(idx == 1)
        )

        fitting = fit_xanes(img, eng_aligned, peakref, color_flag)
        fitting.set_thickness_pnts(len(img))
        fitting.set_thickness()

        if use_full_roi:
            fitting.binary_mask = np.ones(fitting.img[0].shape, dtype=bool)
            fitting.thickness = np.nan_to_num(fitting.thickness)
        else:
            roi_mask_path = os.path.join(roi_mask_dir, f"roi_mask_slice_{idx:04d}.npy")
            if os.path.exists(roi_mask_path):
                roi_mask = np.load(roi_mask_path).astype(bool)
                if roi_mask.shape != fitting.img[0].shape:
                    roi_mask = normalize_mask_to_shape(roi_mask, fitting.img[0].shape)
                fitting.binary_mask = roi_mask
                fitting.thickness = np.nan_to_num(fitting.thickness * roi_mask)
            else:
                fitting.binary_mask = np.ones(fitting.img[0].shape, dtype=bool)
                fitting.thickness = np.nan_to_num(fitting.thickness)

        fitting.polynomial_second_fit_separate(fit_num, ev_step)
        fitting.hsvcolormap(value=3)

        base = os.path.basename(slice_path)
        fn = base[:-4] if base.lower().endswith('.tif') else base[:-5]

        rgb_roi = np.where(fitting.binary_mask[:, :, np.newaxis], fitting.rgb, 0)

        imgs_rgb.append(rgb_roi)
        peak_raw_list.append(fitting.peaksite_raw.copy())
        peaksite_list.append(fitting.peaksite.copy())
        thickness_list.append(fitting.thickness.copy())
        binary_mask_list.append(fitting.binary_mask.copy())
        img_aligned_list.append(fitting.img.copy())
        energy_list.append(eng_aligned.copy())
        slice_names.append(fn)
        slice_indices.append(idx)
        target_shapes.append(fitting.img[0].shape)

    common_h = min(s[0] for s in target_shapes)
    common_w = min(s[1] for s in target_shapes)
    common_shape = (common_h, common_w)

    peak_raw_list = [normalize_array_to_shape(arr, common_shape, fill_value=np.nan) for arr in peak_raw_list]
    peaksite_list = [normalize_array_to_shape(arr, common_shape, fill_value=0) for arr in peaksite_list]
    thickness_list = [normalize_array_to_shape(arr, common_shape, fill_value=0) for arr in thickness_list]
    binary_mask_list = [normalize_mask_to_shape(arr, common_shape) for arr in binary_mask_list]
    imgs_rgb = [normalize_rgb_to_shape(arr, common_shape) for arr in imgs_rgb]
    img_aligned_list = [crop_3d_energy_stack_to_shape(arr, common_shape) for arr in img_aligned_list]
    target_shapes = [common_shape] * len(target_shapes)

    peak_raw_3d = np.stack(peak_raw_list, axis=0).astype(np.float32)
    upper_mask_3d = load_mask_stack(upper_dir, slice_indices, target_shapes) if upper_dir is not None else None
    lower_mask_3d = load_mask_stack(lower_dir, slice_indices, target_shapes) if lower_dir is not None else None
    shell_mask_3d = load_mask_stack(shell_dir, slice_indices, target_shapes) if shell_dir is not None else None
    bulk_mask_3d = load_mask_stack(bulk_dir, slice_indices, target_shapes) if bulk_dir is not None else None

    out_full = os.path.join(parent_dir, f"3D_{n}_fit_{fit_num}★")
    out_upper = os.path.join(parent_dir, f"3D_{n}-1_fit_{fit_num}")
    out_lower = os.path.join(parent_dir, f"3D_{n}-2_fit_{fit_num}")
    out_shell = os.path.join(parent_dir, f"3D_{n}-3_fit_{fit_num}")
    out_bulk = os.path.join(parent_dir, f"3D_{n}-4_fit_{fit_num}")

    full_data_dir = os.path.join(out_full, f"Data_folder_poly2_fit{fit_num}")
    full_hist_dir = os.path.join(full_data_dir, 'histogram')
    full_img_dir = os.path.join(full_data_dir, 'images')
    full_orig_dir = os.path.join(full_img_dir, 'Original')
    full_phase_field_dir = os.path.join(full_img_dir, 'phase_field')
    full_shellbulk_dir = os.path.join(full_img_dir, 'bulk_vs_shell')
    full_h3_shellbulk_dir = os.path.join(full_img_dir, 'H3', 'shell_vs_bulk')
    full_h2_shellbulk_dir = os.path.join(full_img_dir, 'H2', 'shell_vs_bulk')

    # 최종 diverging 이미지는 이 폴더 하나에만 저장됨.
    # 평균 SOC/평균 peak_site 기준 diverging 폴더는 만들지 않음.
    full_diverging_dir = os.path.join(full_img_dir, 'soc_fraction_centered_diverging')
    full_h5_dir = os.path.join(full_data_dir, f"{n}_h5files")

    os.makedirs(full_hist_dir, exist_ok=True)
    os.makedirs(full_orig_dir, exist_ok=True)
    os.makedirs(full_phase_field_dir, exist_ok=True)
    os.makedirs(full_shellbulk_dir, exist_ok=True)
    os.makedirs(full_h3_shellbulk_dir, exist_ok=True)
    os.makedirs(full_h2_shellbulk_dir, exist_ok=True)
    os.makedirs(full_diverging_dir, exist_ok=True)
    os.makedirs(full_h5_dir, exist_ok=True)

    save_rgb_stack(full_orig_dir, slice_names, imgs_rgb)

    full_vals_all = []
    full_energy_raw_all = []

    for raw2d, roi2d in zip(peak_raw_3d, binary_mask_list):
        raw_energy = np.asarray(raw2d[roi2d], dtype=float)
        raw_energy = raw_energy[np.isfinite(raw_energy)]

        energy_filtered = filter_energy_in_range(raw_energy, stats_range[0], stats_range[1])
        if energy_filtered.size:
            full_energy_raw_all.append(energy_filtered)

        vals = compute_stats_single_criterion(raw_energy, stats_range[0], stats_range[1], low_margin, high_margin)
        if vals.size:
            full_vals_all.append(vals)

    full_vals_concat = np.concatenate(full_vals_all) if full_vals_all else np.array([])
    full_energy_raw_concat = np.concatenate(full_energy_raw_all) if full_energy_raw_all else np.array([])

    save_histogram_csv(full_hist_dir, full_vals_concat, stats_range, HIST_BINS)
    save_histogram_raw_csv(full_hist_dir, full_vals_concat, filename='registered_volume_histogram_raw.csv')
    save_peaksite_energy_raw_csv(full_hist_dir, full_energy_raw_concat, filename='peak_site_energy_raw.csv')

    roi_peak_values_for_hist = collect_roi_peak_values(peak_raw_3d, binary_mask_list, stats_range)

    save_probability_histogram_plot_01bin(
        hist_dir=full_hist_dir,
        peak_values=roi_peak_values_for_hist,
        stats_range=stats_range,
        bin_width=0.1
    )

    particle_mean_x = get_particle_mean_peaksite_in_stats_range(peak_raw_3d, binary_mask_list, stats_range)

    # ========================================================
    # 핵심 수정 부분
    # ========================================================
    # 평균 SOC/평균 peak_site 기준 diverging 이미지는 생성하지 않음.
    # 오직 in situ XRD SOC-H3 fraction Excel로 계산한 H2/H3 split energy만
    # diverging white center로 사용한다.
    # ========================================================
    if soc0_peak_site is not None and soc100_peak_site is not None:
        particle_soc = peak_to_soc_percent(
            particle_mean_x,
            soc0_peak_site=soc0_peak_site,
            soc100_peak_site=soc100_peak_site
        )
    else:
        particle_soc = np.nan

    if phase_soc_arr is not None and phase_h3_frac_arr is not None and np.isfinite(particle_soc):
        target_h3_fraction = interpolate_h3_fraction_from_soc(
            particle_soc,
            phase_soc_arr,
            phase_h3_frac_arr
        )
    else:
        target_h3_fraction = np.nan

    h2h3_separation_x = get_separation_energy_from_h3_fraction(
        peak_values=roi_peak_values_for_hist,
        target_h3_fraction=target_h3_fraction,
        stats_range=stats_range
    )

    separation_method = "soc_excel_fraction_quantile"

    if not np.isfinite(h2h3_separation_x):
        raise ValueError(
            "SOC/Excel 기반 H2/H3 separation energy 계산 실패: "
            f"particle_mean_x={particle_mean_x}, "
            f"particle_soc={particle_soc}, "
            f"target_h3_fraction={target_h3_fraction}. "
            "SOC0/SOC100 peak_site, Excel SOC-H3 fraction table, stats_range를 확인하세요."
        )

    print(f"   - particle mean peak_site within stats_range = {particle_mean_x:.6f}")
    print(f"   - SOC0 peak_site = {soc0_peak_site}")
    print(f"   - SOC100 peak_site = {soc100_peak_site}")
    print(f"   - estimated particle SOC = {particle_soc:.6f} %")
    print(f"   - target H3 fraction from Excel = {target_h3_fraction:.6f}")
    print(f"   - H2/H3 split energy for diverging center = {h2h3_separation_x:.6f} ({separation_method})")

    diverging_outline_rgbs = []

    for raw2d, roi2d in zip(peak_raw_3d, binary_mask_list):
        div_rgb = make_diverging_centered_rgb(
            raw2d=raw2d,
            roi2d=roi2d,
            center_x=h2h3_separation_x,
            stats_range=stats_range,
            half_width=0.3
        )

        orange_mask_2d, purple_mask_2d = get_diverging_side_masks(
            raw2d=raw2d,
            roi2d=roi2d,
            center_x=h2h3_separation_x,
            stats_range=stats_range,
            half_width=0.3
        )

        div_outline_rgb = overlay_diverging_domain_centerlines(
            div_rgb=div_rgb,
            orange_mask_2d=orange_mask_2d,
            purple_mask_2d=purple_mask_2d,
            linewidth=0.3
        )

        diverging_outline_rgbs.append(div_outline_rgb)

    # 저장되는 diverging 이미지는 H2/H3 split energy 기준 이미지 하나뿐.
    save_rgb_stack(full_diverging_dir, slice_names, diverging_outline_rgbs)

    save_h5_stack(
        full_h5_dir, slice_names, img_aligned_list, peaksite_list, peak_raw_list,
        thickness_list, imgs_rgb, binary_mask_list, energy_list
    )

    if use_full_roi or missing_masks:
        print(f"✅ registered{n}: 전체 ROI 기준 결과만 저장 완료")
        print(f"   - full  : {out_full}")
        if missing_masks:
            print(f"   - skipped mask outputs because missing: {', '.join(missing_masks)}")
        if np.isfinite(particle_mean_x):
            print(f"   - particle mean peak_site within stats_range = {particle_mean_x:.6f}")
            print(f"   - H2/H3 split energy diverging center = {h2h3_separation_x:.6f}")
        else:
            print("   - particle mean peak_site within stats_range = NaN")
        return

    if upper_mask_3d is not None and lower_mask_3d is not None:
        phase_field_rgbs = []
        for h2_2d, h3_2d in zip(lower_mask_3d, upper_mask_3d):
            phase_field_rgbs.append(make_phase_field_rgb(h2_2d, h3_2d))
        save_rgb_stack(full_phase_field_dir, slice_names, phase_field_rgbs)

    if shell_mask_3d is not None and bulk_mask_3d is not None:
        shellbulk_rgbs = []
        for shell2d, bulk2d in zip(shell_mask_3d, bulk_mask_3d):
            rgb = np.zeros((shell2d.shape[0], shell2d.shape[1], 3), dtype=np.float32)
            rgb[shell2d] = np.array([1.0, 1.0, 0.0], dtype=np.float32)
            rgb[bulk2d] = np.array([1.0, 0.0, 0.0], dtype=np.float32)
            shellbulk_rgbs.append(rgb)
        save_rgb_stack(full_shellbulk_dir, slice_names, shellbulk_rgbs)

        if upper_mask_3d is not None:
            h3_shellbulk_rgbs = []
            for shell2d, bulk2d, h3_2d in zip(shell_mask_3d, bulk_mask_3d, upper_mask_3d):
                h3_shellbulk_rgbs.append(make_core_vs_shell_rgb(shell2d, bulk2d, h3_2d))
            save_rgb_stack(full_h3_shellbulk_dir, slice_names, h3_shellbulk_rgbs)

        if lower_mask_3d is not None:
            h2_shellbulk_rgbs = []
            for shell2d, bulk2d, h2_2d in zip(shell_mask_3d, bulk_mask_3d, lower_mask_3d):
                h2_shellbulk_rgbs.append(make_core_vs_shell_rgb(shell2d, bulk2d, h2_2d))
            save_rgb_stack(full_h2_shellbulk_dir, slice_names, h2_shellbulk_rgbs)

    available_out_roots = []
    if upper_mask_3d is not None:
        available_out_roots.append(out_upper)
    if lower_mask_3d is not None:
        available_out_roots.append(out_lower)
    if shell_mask_3d is not None:
        available_out_roots.append(out_shell)
    if bulk_mask_3d is not None:
        available_out_roots.append(out_bulk)

    for out_root in available_out_roots:
        data_dir = os.path.join(out_root, f"Data_folder_poly2_fit{fit_num}")
        os.makedirs(os.path.join(data_dir, 'histogram'), exist_ok=True)
        os.makedirs(os.path.join(data_dir, 'images'), exist_ok=True)

    for out_root, key_mask in [(out_upper, upper_mask_3d), (out_lower, lower_mask_3d)]:
        if key_mask is None:
            continue
        data_dir = os.path.join(out_root, f"Data_folder_poly2_fit{fit_num}")
        orig_dir = os.path.join(data_dir, 'images', 'Original')
        domain_dir = os.path.join(data_dir, 'images', 'domain_sep')
        core_shell_dir = os.path.join(data_dir, 'images', 'core_vs_shell')
        hist_dir = os.path.join(data_dir, 'histogram')

        masked_rgbs = []
        core_shell_rgbs = []
        vals_all = []
        energy_raw_all = []

        for rgb, raw2d, mask2d, shell2d, bulk2d in zip(imgs_rgb, peak_raw_3d, key_mask, shell_mask_3d, bulk_mask_3d):
            masked_rgbs.append(np.where(mask2d[:, :, np.newaxis], rgb, 0))
            core_shell_rgbs.append(make_core_vs_shell_rgb(shell2d, bulk2d, mask2d))

            raw_energy = np.asarray(raw2d[mask2d], dtype=float)
            raw_energy = raw_energy[np.isfinite(raw_energy)]

            energy_filtered = filter_energy_in_range(raw_energy, stats_range[0], stats_range[1])
            if energy_filtered.size:
                energy_raw_all.append(energy_filtered)

            vals = compute_stats_single_criterion(raw_energy, stats_range[0], stats_range[1], low_margin, high_margin)
            if vals.size:
                vals_all.append(vals)

        vals_concat = np.concatenate(vals_all) if vals_all else np.array([])
        energy_raw_concat = np.concatenate(energy_raw_all) if energy_raw_all else np.array([])

        save_rgb_stack(orig_dir, slice_names, masked_rgbs)
        save_domain_sep_stack(domain_dir, slice_names, key_mask)
        save_rgb_stack(core_shell_dir, slice_names, core_shell_rgbs)
        save_histogram_csv(hist_dir, vals_concat, stats_range, HIST_BINS)
        save_histogram_raw_csv(hist_dir, vals_concat, filename='registered_volume_histogram_raw.csv')
        save_peaksite_energy_raw_csv(hist_dir, energy_raw_concat, filename='peak_site_energy_raw.csv')

    for out_root, region_mask_3d in [(out_shell, shell_mask_3d), (out_bulk, bulk_mask_3d)]:
        if region_mask_3d is None:
            continue

        data_dir = os.path.join(out_root, f"Data_folder_poly2_fit{fit_num}")
        orig_dir = os.path.join(data_dir, 'images', 'Original')
        h3_dir = os.path.join(data_dir, 'images', 'H3')
        h2_dir = os.path.join(data_dir, 'images', 'H2')
        phase_field_dir = os.path.join(data_dir, 'images', 'phase_field')
        hist_dir = os.path.join(data_dir, 'histogram')

        if out_root == out_shell:
            h3_region_color_dir = os.path.join(data_dir, 'images', 'h3_shell')
            h2_region_color_dir = os.path.join(data_dir, 'images', 'h2_shell')
            h3_region_domain_dir = os.path.join(data_dir, 'images', 'h3_shell_domain_sep')
            h2_region_domain_dir = os.path.join(data_dir, 'images', 'h2_shell_domain_sep')
            region_color = [1.0, 1.0, 0.0]
        else:
            h3_region_color_dir = os.path.join(data_dir, 'images', 'h3_bulk')
            h2_region_color_dir = os.path.join(data_dir, 'images', 'h2_bulk')
            h3_region_domain_dir = os.path.join(data_dir, 'images', 'h3_bulk_domain_sep')
            h2_region_domain_dir = os.path.join(data_dir, 'images', 'h2_bulk_domain_sep')
            region_color = [1.0, 0.0, 0.0]

        shell_h3_assigned, bulk_h3_assigned = assign_phase_domains_to_shell_or_bulk(
            upper_mask_3d,
            shell_mask_3d,
            bulk_mask_3d,
            min_connected_voxels=min_connected_voxels_after_combine,
            connectivity=1
        )
        shell_h2_assigned, bulk_h2_assigned = assign_phase_domains_to_shell_or_bulk(
            lower_mask_3d,
            shell_mask_3d,
            bulk_mask_3d,
            min_connected_voxels=min_connected_voxels_after_combine,
            connectivity=1
        )

        if out_root == out_shell:
            region_h3_mask = shell_h3_assigned
            region_h2_mask = shell_h2_assigned
        else:
            region_h3_mask = bulk_h3_assigned
            region_h2_mask = bulk_h2_assigned

        region_original = []
        region_h3 = []
        region_h2 = []
        region_phase_field = []
        region_h3_color_only = []
        region_h2_color_only = []
        vals_all = []
        energy_raw_all = []

        for rgb, raw2d, region2d, h3_2d, h2_2d in zip(
            imgs_rgb, peak_raw_3d, region_mask_3d, region_h3_mask, region_h2_mask
        ):
            region_original.append(np.where(region2d[:, :, np.newaxis], rgb, 0))
            region_h3.append(np.where(h3_2d[:, :, np.newaxis], rgb, 0))
            region_h2.append(np.where(h2_2d[:, :, np.newaxis], rgb, 0))
            region_phase_field.append(make_phase_field_rgb(h2_2d, h3_2d))

            region_h3_color_only.append(make_single_color_mask_rgb(h3_2d, region_color))
            region_h2_color_only.append(make_single_color_mask_rgb(h2_2d, region_color))

            raw_energy = np.asarray(raw2d[region2d], dtype=float)
            raw_energy = raw_energy[np.isfinite(raw_energy)]

            energy_filtered = filter_energy_in_range(raw_energy, stats_range[0], stats_range[1])
            if energy_filtered.size:
                energy_raw_all.append(energy_filtered)

            vals = compute_stats_single_criterion(
                raw_energy, stats_range[0], stats_range[1], low_margin, high_margin
            )
            if vals.size:
                vals_all.append(vals)

        vals_concat = np.concatenate(vals_all) if vals_all else np.array([])
        energy_raw_concat = np.concatenate(energy_raw_all) if energy_raw_all else np.array([])

        save_rgb_stack(orig_dir, slice_names, region_original)
        save_rgb_stack(h3_dir, slice_names, region_h3)
        save_rgb_stack(h2_dir, slice_names, region_h2)
        save_rgb_stack(phase_field_dir, slice_names, region_phase_field)
        save_rgb_stack(h3_region_color_dir, slice_names, region_h3_color_only)
        save_rgb_stack(h2_region_color_dir, slice_names, region_h2_color_only)

        save_domain_sep_stack(h3_region_domain_dir, slice_names, region_h3_mask)
        save_domain_sep_stack(h2_region_domain_dir, slice_names, region_h2_mask)

        save_histogram_csv(hist_dir, vals_concat, stats_range, HIST_BINS)
        save_histogram_raw_csv(hist_dir, vals_concat, filename='registered_volume_histogram_raw.csv')
        save_peaksite_energy_raw_csv(hist_dir, energy_raw_concat, filename='peak_site_energy_raw.csv')

    print(f"✅ registered{n} done")
    print(f"   - full  : {out_full}")
    print(f"   - upper : {out_upper}")
    print(f"   - lower : {out_lower}")
    print(f"   - shell : {out_shell}")
    print(f"   - bulk  : {out_bulk}")
    if np.isfinite(particle_mean_x):
        print(f"   - particle mean peak_site within stats_range = {particle_mean_x:.6f}")
        print(f"   - H2/H3 split energy diverging center = {h2h3_separation_x:.6f}")
    else:
        print("   - particle mean peak_site within stats_range = NaN")


# ============================================================
# Main
# ============================================================

def main():
    root = tk.Tk()
    root.withdraw()

    energy_path = filedialog.askopenfilename(title='energy.txt 선택 (공통)')
    if not energy_path:
        raise ValueError('energy.txt 미선택')

    reverse_stack, reverse_energy = ask_energy_axis_options()
    parents = ask_parent_directories()
    if not parents:
        raise ValueError('상위 폴더 미선택')

    stats_range = ask_stats_range(default_low=STATS_RANGE[0], default_high=STATS_RANGE[1])
    low_margin, high_margin = ask_margins(default_low=0.0, default_high=0.0)

    soc0_peak_site, soc100_peak_site = ask_soc_calibration_peak_sites(
        default_soc0=8353.5,
        default_soc100=8355.2
    )

    phase_excel_path = ask_phase_fraction_excel_path()
    phase_soc_arr, phase_h3_frac_arr = load_soc_h3_fraction_table(phase_excel_path)

    ev_step, peakref = ask_fit_parameters(
        default_ev_step=0.5,
        default_peakref_low=8352.1,
        default_peakref_high=8355.3
    )
    min_connected_voxels_after_combine = ask_min_connected_voxels_after_combine(default_value=3)

    fit_num = 3
    color_flag = 'rgb'
    sli_range = None

    print(f"✅ ev_step = {ev_step}")
    print(f"✅ peakref = ({peakref[0]}, {peakref[1]})")
    print(f"✅ stats_range = ({stats_range[0]}, {stats_range[1]})")
    print("✅ full diverging colormap은 SOC 계산 + Excel 상분율 기반 H2/H3 split energy를 white 기준으로 사용합니다.")
    print(f"✅ SOC calibration: SOC0={soc0_peak_site} eV, SOC100={soc100_peak_site} eV")
    print(f"✅ phase fraction Excel points loaded: {len(phase_soc_arr)}")
    print("✅ histogram 폴더에 0.1 eV binning probability histogram plot도 추가 저장합니다.")
    print("✅ 범위를 넘는 값은 clip 됩니다.")
    print("✅ outline은 orange 영역과 purple 영역 각각의 4-neighbor domain 경계를 사용합니다.")
    print("✅ 검은선 두께는 얇게 linewidth=0.3 로 설정했습니다.")
    print("✅ 평균 SOC/평균 peak_site 기준 diverging 이미지는 저장하지 않습니다.")
    print("✅ soc_fraction_centered_diverging 폴더에는 H2/H3 split energy 기준 최종 이미지만 저장합니다.")

    for parent in parents:
        regs = find_registered_folders(parent)
        if not regs:
            print(f"⚠️ registered 폴더 없음: {parent}")
            continue

        for n, reg_dir in regs:
            try:
                print(f"\n===== processing registered{n}: {reg_dir} =====")
                process_registered_with_saved_masks(
                    parent_dir=parent,
                    reg_dir=reg_dir,
                    n=n,
                    energy_path=energy_path,
                    peakref=peakref,
                    fit_num=fit_num,
                    color_flag=color_flag,
                    sli_range=sli_range,
                    ev_step=ev_step,
                    stats_range=stats_range,
                    low_margin=low_margin,
                    high_margin=high_margin,
                    reverse_stack=reverse_stack,
                    reverse_energy=reverse_energy,
                    min_connected_voxels_after_combine=min_connected_voxels_after_combine,
                    soc0_peak_site=soc0_peak_site,
                    soc100_peak_site=soc100_peak_site,
                    phase_soc_arr=phase_soc_arr,
                    phase_h3_frac_arr=phase_h3_frac_arr
                )
            except Exception as e:
                print(f"❌ registered{n} failed: {e}")

    messagebox.showinfo("완료", "모든 registered 폴더 처리가 완료되었습니다.")


if __name__ == "__main__":
    main()
